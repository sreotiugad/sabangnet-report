import streamlit as st
from google import genai
import math
import os, time, json, hmac, base64, hashlib, traceback, re
import io, zipfile
from urllib.parse import urlparse, parse_qs
from concurrent.futures import ThreadPoolExecutor, as_completed

import numpy as np
import pandas as pd
from datetime import date, timedelta, datetime
from openpyxl import load_workbook
from google.ads.googleads.client import GoogleAdsClient
import requests
from decimal import Decimal, ROUND_HALF_UP, getcontext

getcontext().prec = 28

APP_VERSION = "v2026-01-28-FINAL-ALL-IN-ONE"
# =========================================================
# ✅ 설정: Streamlit Secrets에서 키 로드
# =========================================================

def _secret(key, fallback=""):
    """Streamlit secrets → 환경변수 순으로 조회"""
    try:
        return st.secrets[key]
    except Exception:
        return os.environ.get(key, fallback)

# =========================================================
# Gemini
# =========================================================
GEMINI_API_KEY = _secret("GEMINI_API_KEY")
GEMINI_MODEL   = "gemini-2.5-flash"
_gemini = genai.Client(api_key=GEMINI_API_KEY) if GEMINI_API_KEY else None

def _raise_no_key():
    raise RuntimeError('GEMINI_API_KEY가 설정되어 있지 않습니다. Streamlit secrets 또는 환경변수로 설정하세요.')


# ---- Google Ads ----
os.environ["GADS_DEVELOPER_TOKEN"]  = _secret("GADS_DEVELOPER_TOKEN")
os.environ["GADS_CLIENT_ID"]        = _secret("GADS_CLIENT_ID")
os.environ["GADS_CLIENT_SECRET"]    = _secret("GADS_CLIENT_SECRET")
os.environ["GADS_REFRESH_TOKEN"]    = _secret("GADS_REFRESH_TOKEN")
os.environ["GADS_LOGIN_CUSTOMER_ID"]= _secret("GADS_LOGIN_CUSTOMER_ID")
os.environ["GADS_CUSTOMER_ID"]      = _secret("GADS_CUSTOMER_ID")
os.environ["GADS_CUSTOMER_ID_2"]    = _secret("GADS_CUSTOMER_ID_2")

GOOGLE_CUSTOMER_IDS = [
    os.environ["GADS_CUSTOMER_ID"],
    os.environ["GADS_CUSTOMER_ID_2"],
]

# ---- Naver SearchAd (2계정) ----
os.environ["NAVER1_CUSTOMER_ID"] = _secret("NAVER1_CUSTOMER_ID")
os.environ["NAVER1_API_KEY"]     = _secret("NAVER1_API_KEY")
os.environ["NAVER1_SECRET_KEY"]  = _secret("NAVER1_SECRET_KEY")

os.environ["NAVER2_CUSTOMER_ID"] = _secret("NAVER2_CUSTOMER_ID")
os.environ["NAVER2_API_KEY"]     = _secret("NAVER2_API_KEY")
os.environ["NAVER2_SECRET_KEY"]  = _secret("NAVER2_SECRET_KEY")



NAVER_BASE_URL = "https://api.searchad.naver.com"
MAX_CAMPAIGNS_PER_ACCOUNT = 500

def _load_naver_accounts():
    accs = []
    for i in [1, 2]:
        cid = os.environ.get(f"NAVER{i}_CUSTOMER_ID")
        key = os.environ.get(f"NAVER{i}_API_KEY")
        sec = os.environ.get(f"NAVER{i}_SECRET_KEY")
        if cid and key and sec:
            accs.append({"customer_id": cid, "api_key": key, "secret_key": sec})
    return accs

NAVER_ACCOUNTS = _load_naver_accounts()

# =========================================================
# ✅ 1) 브랜드검색 일별 광고비 (VAT 포함)
# =========================================================
BS_DAILY_FEE_VAT_INCLUDED = {
    "사방넷_BS_MO": Decimal("2640000") / Decimal("90"),
    "사방넷_BS_PC": Decimal("3960000") / Decimal("90"),
    "풀필먼트_BS_PC": Decimal("1980000") / Decimal("90"),
    "풀필먼트_BS_MO": Decimal("1980000") / Decimal("90"),
    "미니_BS_PC": Decimal("1980000") / Decimal("90"),
    "미니_BS_MO": Decimal("2640000") / Decimal("90"),
}

# =========================================================
# ✅ 2) 공용 유틸(절대 안전)
# =========================================================
def round_half_up_int(x) -> int:
    try:
        d = Decimal(str(x))
        return int(d.quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    except Exception:
        return 0

def col_or_zeros(df: pd.DataFrame, col: str, n: int) -> pd.Series:
    """df에 col이 없으면 길이 n짜리 0 시리즈 반환"""
    if col in df.columns:
        return pd.to_numeric(df[col], errors="coerce").fillna(0)
    return pd.Series([0] * n)

def yyyymmdd(iso_date: str) -> str:
    return str(iso_date).replace("-", "")[:8]

def assign_service_from_campaign(campaign_series: pd.Series) -> pd.Series:
    s = campaign_series.astype(str)
    return np.select(
        [
            s.str.contains("풀필먼트", na=False),
            s.str.contains("미니", na=False),
            s.str.contains("사방넷", na=False),
        ],
        ["풀필먼트", "사방넷미니", "사방넷"],
        default=""
    )

def calc_display_cost(row) -> int:
    """
    광고비(마크업포함,VAT포함)
    - 구글: 총비용(VAT제외) * 1.1 -> VAT포함
    - 네이버: 총비용(salesAmt, 이미 VAT포함) / 1.1 -> VAT제외 기준 정산액
    ✅ 수정: 네이버 * 1.1 → / 1.1 (salesAmt는 이미 VAT포함이므로)
    """
    cost = Decimal(str(row.get("총비용", 0) or 0))
    media = str(row.get("매체", ""))

    if media == "구글":
        val = cost * Decimal("1.1")
    elif media == "네이버":
        val = cost / Decimal("1.1")   # ✅ FIX: * → /
    else:
        val = cost

    return int(val.quantize(Decimal("1"), rounding=ROUND_HALF_UP))

# =========================================================
# ✅ 3) 날짜/캘린더
# =========================================================

def preset_range(preset: str):
    today = date.today()

    if preset == "주간(월~일)":
        end = today
        monday = end - timedelta(days=end.weekday())
        sunday = monday + timedelta(days=6)
        s, e = monday, sunday
    elif preset == "어제":
        s = e = today - timedelta(days=1)
    elif preset == "지난 7일":
        e = today - timedelta(days=1)
        s = e - timedelta(days=6)
    elif preset == "지난 30일":
        e = today - timedelta(days=1)
        s = e - timedelta(days=29)
    elif preset == "이번 달":
        e = today - timedelta(days=1)
        s = e.replace(day=1)
    else:
        s, e = today - timedelta(days=6), today

    return s.isoformat(), e.isoformat()


WEEKDAY_KO = {0:"월",1:"화",2:"수",3:"목",4:"금",5:"토",6:"일"}

def add_cal_fields(df, date_col="날짜"):
    df = df.copy()

    raw = df[date_col].astype(str).str.strip()
    raw = raw.replace({"None": "", "nan": "", "NaT": ""})

    raw2 = raw.copy()
    m = raw2.str.fullmatch(r"\d{8}", na=False)
    raw2.loc[m] = raw2.loc[m].str.slice(0,4) + "-" + raw2.loc[m].str.slice(4,6) + "-" + raw2.loc[m].str.slice(6,8)

    parsed = pd.to_datetime(raw2, errors="coerce")
    df["기간"] = parsed
    mask = parsed.notna()

    df["요일"] = ""
    df["year"] = pd.NA
    df["month"] = ""
    df["week"] = pd.NA
    df["week시작"] = pd.NaT
    df["week종료"] = pd.NaT

    df.loc[mask, "요일"] = parsed.loc[mask].dt.weekday.map(WEEKDAY_KO)
    df.loc[mask, "year"] = parsed.loc[mask].dt.year
    df.loc[mask, "month"] = parsed.loc[mask].dt.year.astype(str) + "." + parsed.loc[mask].dt.month.astype(str)

    d = parsed.loc[mask]
    ws = d - pd.to_timedelta(d.dt.weekday, unit="D")
    we = ws + pd.to_timedelta(6, unit="D")

    ms = d.dt.to_period("M").dt.start_time
    me = d.dt.to_period("M").dt.end_time

    df.loc[mask, "week시작"] = np.where(ws < ms, ms, ws)
    df.loc[mask, "week종료"] = np.where(we > me, me, we)

    iso = parsed.loc[mask].dt.isocalendar()
    df.loc[mask, "week"] = iso["week"].astype("Int64").values
    df.loc[mask, "year"] = iso["year"].astype("Int64").values

    return df

# =========================================================
# ✅ 4) Naver 공용 (서명/다운로드)
# =========================================================
# ✅ 4) Naver 공용 (서명/다운로드) - SINGLE SOURCE OF TRUTH
# =========================================================

def naver_headers(acc, uri: str, method: str = "GET") -> dict:
    ts = str(int(time.time() * 1000))

    method = method.upper().strip()
    uri = str(uri).strip()

    msg = f"{ts}.{method}.{uri}"

    # ✅ secretKey는 base64 decode 하지 말고 '문자열 그대로' 사용
    secret = str(acc["secret_key"]).strip().encode("utf-8")

    sig = base64.b64encode(
        hmac.new(secret, msg.encode("utf-8"), hashlib.sha256).digest()
    ).decode()

    return {
        "X-Timestamp": ts,
        "X-API-KEY": str(acc["api_key"]).strip(),
        "X-Customer": str(acc["customer_id"]).strip(),
        "X-Signature": sig,
        "Content-Type": "application/json",
    }

def _split_download_url(download_url: str):
    """
    download_url이
      - '/report-download?authtoken=...&fileVersion=v2' (상대경로)
      - 'https://api.searchad.naver.com/report-download?....' (절대경로)
    어떤 형태로 오든,
    path('/report-download') 와 params(dict) 로 분리해준다.
    """
    full = download_url if download_url.startswith("http") else (NAVER_BASE_URL + download_url)
    u = urlparse(full)

    path = u.path                    # ✅ '/report-download'
    qs = parse_qs(u.query)

    # parse_qs는 값이 리스트라서 첫 값만 꺼냄
    params = {k: v[0] for k, v in qs.items() if v}

    return path, params


def naver_download_report(acc, download_url: str) -> bytes:
    """
    ✅ 핵심 규칙:
    - 서명에 넣는 uri = path만 사용 (쿼리스트링 제외)  -> '/report-download'
    - 실제 요청은 params로 authtoken/fileVersion 전달
    """
    path, params = _split_download_url(download_url)

    url = NAVER_BASE_URL + path

    r = requests.get(
        url,
        params=params,
        headers=naver_headers(acc, path, "GET"),   # ✅ 서명은 path만!
        timeout=120
    )

    if r.status_code != 200:
        raise Exception(f"NAVER report-download 실패 status={r.status_code} body={r.text[:300]}")

    return r.content



def safe_json(resp):
    try:
        return resp.json()
    except Exception:
        return None

def normalize_naver_date(raw_date):
    if raw_date is None:
        return None
    s = str(raw_date).strip().rstrip(".")
    if re.fullmatch(r"\d{8}", s):
        return f"{s[:4]}-{s[4:6]}-{s[6:8]}"
    if re.fullmatch(r"\d{4}\.\d{2}\.\d{2}", s):
        y, m, d = s.split(".")
        return f"{y}-{m}-{d}"
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        return s[:10]
    return None

def pick_naver_date_from_item(item: dict):
    candidates = [
        item.get("date"), item.get("statDt"), item.get("statDate"),
        item.get("stat_date"), item.get("dt"), item.get("day"), item.get("ymd"),
    ]
    for v in candidates:
        dt = normalize_naver_date(v)
        if dt and re.fullmatch(r"\d{4}-\d{2}-\d{2}", dt):
            return dt
    for _, v in item.items():
        dt = normalize_naver_date(v)
        if dt and re.fullmatch(r"\d{4}-\d{2}-\d{2}", dt):
            return dt
    return None

def pick_naver_device_from_item(item: dict):
    raw = item.get("pcMblTp") or item.get("device") or ""
    if raw == "PC":
        return "PC"
    if raw in ["MOBILE", "MOB", "모바일", "M"]:
        return "모바일"
    return ""

def infer_device_from_campaign_name(cname: str) -> str:
    s = str(cname or "")
    if s.endswith("_PC") or s.endswith("PC"):
        return "PC"
    if s.endswith("_MO") or s.endswith("MO") or s.endswith("_M") or "모바일" in s:
        return "모바일"
    return "전체"

def naver_list_campaigns(acc):
    uri = "/ncc/campaigns"
    r = requests.get(NAVER_BASE_URL + uri, headers=naver_headers(acc, uri, "GET"), timeout=30)

    if r.status_code != 200:
        raise Exception(f"[NAVER] /ncc/campaigns 실패 status={r.status_code} body={r.text[:300]}")

    j = safe_json(r)
    return j if isinstance(j, list) else []

def naver_list_adgroups(acc, campaign_id: str = None):
    uri = "/ncc/adgroups"
    params = {}
    if campaign_id:
        params["nccCampaignId"] = campaign_id

    r = requests.get(
        NAVER_BASE_URL + uri,
        headers=naver_headers(acc, uri, "GET"),
        params=params,
        timeout=30
    )
    if r.status_code != 200:
        raise Exception(f"[NAVER] /ncc/adgroups 실패 status={r.status_code} body={r.text[:300]}")
    j = safe_json(r)
    return j if isinstance(j, list) else []

def _naver_master_report(acc, item: str, logs=None) -> list:
    """
    Master Report API로 특정 item 전체 목록을 한 번에 가져온다.
    POST /master-reports → 폴링 → 다운로드 → JSON 파싱
    """
    if logs is None:
        logs = []

    uri = "/master-reports"
    r = requests.post(
        NAVER_BASE_URL + uri,
        headers=naver_headers(acc, uri, "POST"),
        json={"item": item},
        timeout=30,
    )
    if r.status_code not in (200, 201):
        logs.append(f"❌ [MasterReport] {item} 생성 실패 status={r.status_code} body={r.text[:200]}")
        return []

    job = r.json()
    job_id = job.get("id")
    if not job_id:
        logs.append(f"❌ [MasterReport] {item} job_id 없음 resp={str(job)[:200]}")
        return []

    # 폴링
    download_url = None
    for i in range(30):
        uri_status = f"/master-reports/{job_id}"
        rs = requests.get(
            NAVER_BASE_URL + uri_status,
            headers=naver_headers(acc, uri_status, "GET"),
            timeout=30,
        )
        if rs.status_code != 200:
            time.sleep(2)
            continue
        st = rs.json()
        status = str(st.get("status", "")).upper()
        du = st.get("downloadUrl") or st.get("downloadURL")
        logs.append(f"[MasterReport] {item} poll={i+1} status={status}")
        if status == "BUILT" and du:
            download_url = du
            break
        if status in ("ERROR", "NONE"):
            logs.append(f"❌ [MasterReport] {item} 빌드 실패 status={status}")
            return []
        time.sleep(2)

    if not download_url:
        logs.append(f"⚠️ [MasterReport] {item} downloadUrl 없음 (timeout)")
        return []

    # 다운로드
    try:
        content = naver_download_report(acc, download_url)
    except Exception as e:
        logs.append(f"❌ [MasterReport] {item} 다운로드 실패: {e}")
        return []

    # ZIP 해제
    if content[:2] == b"PK":
        z = zipfile.ZipFile(io.BytesIO(content))
        content = z.read(z.namelist()[0])

    # JSON 파싱
    try:
        data = json.loads(content.decode("utf-8"))
        if isinstance(data, list):
            return data
        return data.get("items", data.get("data", []))
    except Exception as e:
        logs.append(f"❌ [MasterReport] {item} JSON 파싱 실패: {e} / 원본앞: {content[:200]}")
        return []


def naver_build_name_maps(acc, exclude_bs=False, logs=None):
    """
    Master Report API로 캠페인/그룹/키워드 id→name 맵을 한 번에 구성.
    기존: 캠페인 N번 + 그룹 N번 + 키워드 N번 API 호출 (수백 번)
    변경: Campaign / Adgroup / Keyword 각 1번 호출 (총 3번)
    """
    if logs is None:
        logs = []

    # ── 캠페인 맵 ──────────────────────────────────────────
    camp_items = _naver_master_report(acc, "Campaign", logs)
    camp_map = {}
    for c in camp_items:
        cid  = c.get("nccCampaignId") or c.get("campaignId") or c.get("id")
        name = c.get("name") or c.get("campaignName")
        if cid and name:
            camp_map[cid] = name

    # BS 제외 (키워드 리포트용)
    if exclude_bs:
        camp_map = {
            cid: name for cid, name in camp_map.items()
            if "_BS_" not in str(name) and "-a001-04-" not in str(cid)
        }

    logs.append(f"[MasterReport] Campaign 맵: {len(camp_map)}개")

    # ── 그룹 맵 ──────────────────────────────────────────
    grp_items = _naver_master_report(acc, "Adgroup", logs)
    grp_map = {}
    for g in grp_items:
        gid  = g.get("nccAdgroupId") or g.get("adgroupId") or g.get("id")
        name = g.get("name") or g.get("adgroupName")
        if gid and name:
            grp_map[gid] = name

    logs.append(f"[MasterReport] Adgroup 맵: {len(grp_map)}개")

    # ── 키워드 맵 ──────────────────────────────────────────
    kw_items = _naver_master_report(acc, "Keyword", logs)
    kw_map = {}
    for k in kw_items:
        kid  = k.get("nccKeywordId") or k.get("keywordId") or k.get("id")
        name = k.get("keyword") or k.get("keywordName") or k.get("name")
        # BS 캠페인 소속 키워드 제외
        if exclude_bs:
            camp_id = k.get("nccCampaignId") or k.get("campaignId") or ""
            if "-a001-04-" in str(camp_id):
                continue
        if kid and name:
            kw_map[kid] = name

    logs.append(f"[MasterReport] Keyword 맵: {len(kw_map)}개")

    return camp_map, grp_map, kw_map

def naver_fetch_stats_by_id(acc, cid, since_yyyymmdd, until_yyyymmdd, breakdown=True):
    uri = "/stats"
    params = {
        "id": cid,
        "fields": json.dumps(["impCnt", "clkCnt", "salesAmt", "ccnt"]),
        "timeRange": json.dumps({"since": since_yyyymmdd, "until": until_yyyymmdd}),
        "timeIncrement": "1",
    }
    if breakdown:
        params["breakdown"] = "pcMblTp"

    return requests.get(
        NAVER_BASE_URL + uri,
        headers=naver_headers(acc, uri, "GET"),
        params=params,
        timeout=60,
    )

def _date_list_yyyymmdd(d_from: str, d_to: str):
    s = datetime.strptime(d_from[:10], "%Y-%m-%d").date()
    e = datetime.strptime(d_to[:10], "%Y-%m-%d").date()
    out = []
    cur = s
    while cur <= e:
        out.append(cur.strftime("%Y%m%d"))
        cur += timedelta(days=1)
    return out

def naver_statdt(x) -> str:
    s = str(x).strip().rstrip(".")
    s = s.replace(".", "").replace("-", "")
    return s[:8]

def naver_create_stat_report(acc, report_tp: str, stat_yyyy_mm_dd: str, stat_level: str = "KEYWORD"):
    uri = "/stat-reports"
    payload = {
        "reportTp": report_tp,
        "statDt": naver_statdt(stat_yyyy_mm_dd),
        "statLevel": stat_level
    }

    r = requests.post(
        NAVER_BASE_URL + uri,
        headers=naver_headers(acc, uri, "POST"),
        json=payload,
        timeout=30
    )

    # ✅ 200뿐 아니라 201도 성공으로 처리
    if r.status_code not in (200, 201):
        raise Exception(
            "NAVER /stat-reports 요청 실패\n"
            f"status_code: {r.status_code}\n"
            f"payload: {payload}\n"
            f"response_text: {r.text}"
        )

    return r.json()

def naver_get_stat_report_status(acc, report_job_id: str):
    uri = f"/stat-reports/{report_job_id}"
    r = requests.get(NAVER_BASE_URL + uri, headers=naver_headers(acc, uri, "GET"), timeout=30)
    r.raise_for_status()
    return r.json()




def get_n_keyword_data_report(d_from, d_to, report_tp="AD", logs=None) -> pd.DataFrame:
    if logs is None:
        logs = []

    all_dfs = []
    days = _date_list_yyyymmdd(d_from, d_to)

    for acc in NAVER_ACCOUNTS:
        logs.append(f"[NAVER] account customer_id={acc.get('customer_id')} reportTp={report_tp}")
        camp_map, grp_map, kw_map = naver_build_name_maps(acc, exclude_bs=True, logs=logs)
        logs.append(f"[NAVER] name_maps: camp={len(camp_map)}, grp={len(grp_map)}, kw={len(kw_map)}")

        for day in days:
            # 노출/클릭/비용 리포트
            df_ad = _fetch_naver_report_day(acc, day, report_tp, camp_map, grp_map, kw_map, logs)
            if df_ad is None:
                continue

            # ✅ BS 캠페인 행 제거
            before = len(df_ad)
            df_ad = df_ad[~df_ad["campaignId"].astype(str).str.contains("-a001-04-", na=False)].reset_index(drop=True)
            if len(df_ad) < before:
                logs.append(f"[NAVER] BS 캠페인 행 제거: {before}→{len(df_ad)} day={day}")

            # ✅ adId 집계를 먼저! (ccnt 머지 전에 해야 중복 방지)
            if "keywordId" in df_ad.columns:
                grp_cols = [c for c in ["statDt","customerId","campaignId","adgroupId","keywordId","pcMblTp","campaignName","adgroupName","keywordName"] if c in df_ad.columns]
                sum_cols = [c for c in ["impCnt","clkCnt","clkAmt"] if c in df_ad.columns]
                agg_dict = {c: "sum" for c in sum_cols}
                if "avgRnk" in df_ad.columns:
                    agg_dict["avgRnk"] = "mean"
                if "bidAmt" in df_ad.columns:
                    agg_dict["bidAmt"] = "first"
                df_ad = df_ad.groupby(grp_cols, as_index=False).agg(agg_dict)

            # ✅ AD_CONVERSION 머지 (집계 후에 ccnt 붙이기)
            if report_tp != "EXPKEYWORD":
                df_conv = _fetch_naver_report_day(acc, day, "AD_CONVERSION", camp_map, grp_map, kw_map, logs)
            else:
                df_conv = None

            if df_conv is not None and "ccnt" in df_conv.columns:
                df_conv = df_conv[~df_conv["campaignId"].astype(str).str.contains("-a001-04-", na=False)]

                conv_kw = df_conv[df_conv["keywordId"].astype(str).str.strip() != "-"]
                conv_kw_agg = conv_kw.groupby(["keywordId","pcMblTp"], as_index=False)["ccnt"].sum()
                df_ad = df_ad.merge(conv_kw_agg, on=["keywordId","pcMblTp"], how="left")
                df_ad["ccnt"] = df_ad["ccnt"].fillna(0)

                conv_grp = df_conv[df_conv["keywordId"].astype(str).str.strip() == "-"]
                if not conv_grp.empty:
                    conv_grp_agg = conv_grp.groupby(["adgroupId","pcMblTp"], as_index=False)["ccnt"].sum().rename(columns={"ccnt":"ccnt_grp"})
                    df_ad = df_ad.merge(conv_grp_agg, on=["adgroupId","pcMblTp"], how="left")
                    mask = df_ad["keywordId"].astype(str).str.strip() == "-"
                    df_ad.loc[mask, "ccnt"] = df_ad.loc[mask, "ccnt_grp"].fillna(0)
                    df_ad.drop(columns=["ccnt_grp"], inplace=True)

                logs.append(f"[NAVER] AD+CONVERSION 머지 완료 day={day}")
            else:
                df_ad["ccnt"] = 0

            all_dfs.append(df_ad)

    if not all_dfs:
        return pd.DataFrame()

    result = pd.concat(all_dfs, ignore_index=True)

    if result.empty:
        return result

    # 중복 제거 (날짜별 집계는 루프 내에서 이미 완료)
    dedup_cols = ["statDt","campaignId","adgroupId","keywordName","pcMblTp"]
    if "keywordId" in result.columns:
        dedup_cols = ["statDt","campaignId","adgroupId","keywordId","pcMblTp"]
    existing_dedup = [c for c in dedup_cols if c in result.columns]
    before = len(result)
    result = result.drop_duplicates(subset=existing_dedup).reset_index(drop=True)
    after = len(result)
    if before != after:
        logs.append(f"[NAVER] 중복 제거: {before}행 → {after}행")

    return result


def _fetch_naver_report_day(acc, day, report_tp, camp_map, grp_map, kw_map, logs):
    """단일 날짜 네이버 리포트 다운로드 → DataFrame 반환 (실패 시 None)"""
    try:
        job = naver_create_stat_report(acc, report_tp=report_tp, stat_yyyy_mm_dd=day)
    except Exception as e:
        logs.append(f"❌ create_stat_report failed day={day} reportTp={report_tp} err={e}")
        return None

    job_id = job.get("reportJobId") or job.get("reportJobID") or job.get("reportId")
    if not job_id:
        logs.append(f"❌ job_id missing day={day} reportTp={report_tp} resp={str(job)[:200]}")
        return None

    download_url = None
    last_status = None
    for i in range(30):
        st = naver_get_stat_report_status(acc, str(job_id))
        status = str(st.get("status", "")).upper()
        du = st.get("downloadUrl") or st.get("downloadURL") or st.get("download_url")
        if status != last_status:
            logs.append(f"[NAVER] day={day} reportTp={report_tp} poll={i+1} status={status} downloadUrl={'Y' if du else 'N'}")
            last_status = status
        if status in ("BUILT", "DONE", "COMPLETED", "SUCCESS") and du:
            download_url = du
            break
        if status in ("ERROR", "FAIL", "FAILED"):
            logs.append(f"❌ report build failed day={day} reportTp={report_tp} status={status}")
            return None
        time.sleep(2)

    if not download_url:
        logs.append(f"⚠️ no downloadUrl day={day} reportTp={report_tp} (skipped)")
        return None

    try:
        content = naver_download_report(acc, download_url)
        logs.append(f"[NAVER] downloaded day={day} reportTp={report_tp} bytes={len(content)}")
    except Exception as e:
        logs.append(f"❌ download failed day={day} reportTp={report_tp} err={e}")
        return None

    if content[:2] == b"PK":
        z = zipfile.ZipFile(io.BytesIO(content))
        csv_bytes = z.read(z.namelist()[0])
    else:
        csv_bytes = content

    try:
        try:
            txt = csv_bytes.decode("utf-8")
        except UnicodeDecodeError:
            txt = csv_bytes.decode("cp949", errors="replace")

        first_line = txt.splitlines()[0] if txt.strip() else ""
        col_count = len(first_line.split("\t"))
        logs.append(f"[NAVER] col_count={col_count} reportTp={report_tp}")
        # ✅ AD_CONVERSION raw 확인용
        if report_tp == "AD_CONVERSION":
            head3 = "\n".join(txt.splitlines()[:3])
            logs.append(f"[NAVER] AD_CONVERSION raw:\n{head3}")
        if report_tp == "AD":
            head2 = "\n".join(txt.splitlines()[:2])
            logs.append(f"[NAVER] AD raw 샘플:\n{head2}")
        if report_tp == "EXPKEYWORD":
            head2 = "\n".join(txt.splitlines()[:2])
            logs.append(f"[NAVER] EXPKEYWORD raw 샘플:\n{head2}")

        if report_tp == "AD":
            base_cols = [
                "statDt","customerId","campaignId","adgroupId",
                "keywordId","adId","bsnId","bidAmt","pcMblTp",
                "impCnt","clkCnt","clkAmt","convAmt","avgRnk"  # clkAmt=광고비, convAmt=전환매출
            ]
            if col_count >= 15:
                base_cols.append("cpConv")
        elif report_tp == "EXPKEYWORD":
            # 12컬럼: statDt customerId campaignId adgroupId keywordName bidAmt pcMblTp impCnt clkCnt ccnt salesAmt avgRnk
            base_cols = [
                "statDt","customerId","campaignId","adgroupId",
                "keywordName","bidAmt","pcMblTp",
                "impCnt","clkCnt","ccnt","salesAmt","avgRnk"
            ]
        else:  # AD_CONVERSION - 실제 13컬럼
            # statDt, customerId, campaignId, adgroupId, keywordId, adId, bsnId, bidAmt, pcMblTp, clkCnt, convType, ccnt, (미확인)
            base_cols = [
                "statDt","customerId","campaignId","adgroupId",
                "keywordId","adId","bsnId","bidAmt","pcMblTp",
                "clkCnt","convType","ccnt","extra1"
            ]

        df = pd.read_csv(io.StringIO(txt), sep="\t", header=None, names=base_cols, engine="python")
        df["campaignName"] = df["campaignId"].map(camp_map).fillna(df["campaignId"])
        df["adgroupName"]  = df["adgroupId"].map(grp_map).fillna(df["adgroupId"])
        if report_tp == "EXPKEYWORD":
            # keywordName이 이미 컬럼에 있음, keywordId는 없음
            if "keywordId" not in df.columns:
                df["keywordId"] = df["keywordName"]
        else:
            df["keywordName"] = df["keywordId"].map(kw_map).fillna(df["keywordId"])
        if "statDt" not in df.columns:
            df["statDt"] = day
        logs.append(f"[NAVER] parsed rows day={day} reportTp={report_tp} rows={len(df)}")
        return df

    except Exception as e:
        logs.append(f"❌ csv parse failed day={day} reportTp={report_tp} err={e}")
        return None

# =========================================================
# ✅ 5) Naver 일반 리포트(캠페인단위)
# =========================================================
def _date_range_iso(d_from: str, d_to: str):
    s = datetime.strptime(d_from[:10], "%Y-%m-%d").date()
    e = datetime.strptime(d_to[:10], "%Y-%m-%d").date()
    out = []
    cur = s
    while cur <= e:
        out.append(cur.isoformat())
        cur += timedelta(days=1)
    return out

def fill_missing_brandsearch_rows(df: pd.DataFrame, d_from: str, d_to: str) -> pd.DataFrame:
    bs_keys = list(BS_DAILY_FEE_VAT_INCLUDED.keys())
    all_dates = _date_range_iso(d_from, d_to)

    if df is None or df.empty:
        df = pd.DataFrame(columns=["매체구분","매체","캠페인유형","캠페인","날짜","기기","노출수","클릭수","총비용","가입"])

    existing = set(zip(df["날짜"].astype(str), df["캠페인"].astype(str)))

    new_rows = []
    for dt in all_dates:
        for camp in bs_keys:
            if (dt, camp) in existing:
                continue
            device = "PC" if camp.endswith("_PC") else "모바일"
            fee = BS_DAILY_FEE_VAT_INCLUDED[camp]
            new_rows.append({
                "매체구분": "SA",
                "매체": "네이버",
                "캠페인유형": "브랜드검색/신제품검색",
                "캠페인": camp,
                "날짜": dt,
                "기기": device,
                "노출수": 0,
                "클릭수": 0,
                "총비용": float(fee),
                "가입": 0.0,
            })

    if new_rows:
        df = pd.concat([df, pd.DataFrame(new_rows)], ignore_index=True)

    is_bs = df["캠페인"].isin(bs_keys)
    if is_bs.any():
        df.loc[is_bs, "총비용"] = df.loc[is_bs, "캠페인"].map(lambda k: float(BS_DAILY_FEE_VAT_INCLUDED[k]))
        df.loc[is_bs, "기기"] = np.where(df.loc[is_bs, "캠페인"].astype(str).str.endswith("_PC"), "PC", "모바일")
        df.loc[is_bs, "캠페인유형"] = "브랜드검색/신제품검색"

    return df

def get_n_data(d_from, d_to, logs=None):
    if logs is None:
        logs = []

    rows = []
    since = yyyymmdd(d_from)
    until = yyyymmdd(d_to)

    NAVER_CAMPAIGN_TP_MAP = {"WEB_SITE": "파워링크"}
    bs_keys = set(BS_DAILY_FEE_VAT_INCLUDED.keys())

    for acc in NAVER_ACCOUNTS:
        camps = naver_list_campaigns(acc)[:MAX_CAMPAIGNS_PER_ACCOUNT]

        for camp in camps:
            cid = camp.get("nccCampaignId")
            cname = camp.get("name", cid)

            if not cid:
                continue

            r = naver_fetch_stats_by_id(acc, cid, since, until, breakdown=True)

            if r.status_code == 400:
                r = naver_fetch_stats_by_id(acc, cid, since, until, breakdown=False)

            if r.status_code != 200:
                logs.append(f"❌ NAVER /stats 실패 cname={cname} status={r.status_code} body={r.text[:200]}")
                continue

            sj = safe_json(r) or {}
            data = sj.get("data", [])

            if not data:
                continue

            tp_raw = str(camp.get("campaignTp", "SA") or "SA")
            tp_fix = NAVER_CAMPAIGN_TP_MAP.get(tp_raw, tp_raw)

            for item in data:
                dt_norm = pick_naver_date_from_item(item)
                if not dt_norm:
                    continue

                device = pick_naver_device_from_item(item)
                if not device:
                    device = infer_device_from_campaign_name(cname)

                imp = int(item.get("impCnt", 0) or 0)
                clk = int(item.get("clkCnt", 0) or 0)
                conv = float(item.get("ccnt", 0) or 0)
                cost = Decimal(str(item.get("salesAmt", 0) or 0))

                if cname in bs_keys:
                    cost = BS_DAILY_FEE_VAT_INCLUDED.get(cname, cost)

                if (imp == 0) and (cname not in bs_keys):
                    continue

                rows.append({
                    "매체구분": "SA",
                    "매체": "네이버",
                    "캠페인유형": "브랜드검색/신제품검색" if cname in bs_keys else tp_fix,
                    "캠페인": cname,
                    "날짜": dt_norm,
                    "기기": device,
                    "노출수": imp,
                    "클릭수": clk,
                    "총비용": float(cost),
                    "가입": conv,
                })

    df = pd.DataFrame(rows)
    df = fill_missing_brandsearch_rows(df, d_from, d_to)
    return df, logs



# =========================================================
# ✅ 6) Google 일반/키워드
# =========================================================
def _google_client():
    cfg = {
        "developer_token": os.environ["GADS_DEVELOPER_TOKEN"],
        "client_id": os.environ["GADS_CLIENT_ID"],
        "client_secret": os.environ["GADS_CLIENT_SECRET"],
        "refresh_token": os.environ["GADS_REFRESH_TOKEN"],
        "use_proto_plus": True,
        "login_customer_id": os.environ.get("GADS_LOGIN_CUSTOMER_ID"),
    }
    return GoogleAdsClient.load_from_dict(cfg)

def get_g_data(d_from, d_to, logs=None):
    if logs is None:
        logs = []

    client = _google_client()
    ga = client.get_service("GoogleAdsService")

    query = f"""
        SELECT
          segments.date,
          campaign.advertising_channel_type,
          campaign.name,
          segments.device,
          metrics.impressions,
          metrics.clicks,
          metrics.cost_micros,
          metrics.conversions
        FROM campaign
        WHERE segments.date BETWEEN '{d_from}' AND '{d_to}'
    """.strip()

    rows = []
    d_map = {"DESKTOP": "PC", "MOBILE": "모바일", "TABLET": "모바일"}
    div_map = {"SEARCH": "SA", "DISPLAY": "DA", "VIDEO": "VA", "PERFORMANCE_MAX": "DA"}
    type_ko = {"SEARCH": "검색", "DISPLAY": "디스플레이", "VIDEO": "동영상", "PERFORMANCE_MAX": "실적 최대화"}

    # ✅ 하위 광고주 2개 모두 조회
    for cust_id in GOOGLE_CUSTOMER_IDS:
        try:
            logs.append(f"[Google] customer_id={cust_id} 조회 시작")
            stream = ga.search_stream(customer_id=cust_id, query=query)
            count = 0
            for b in stream:
                for r in b.results:
                    ch = r.campaign.advertising_channel_type.name
                    rows.append({
                        "매체구분": div_map.get(ch, "SA"),
                        "매체": "구글",
                        "캠페인유형": type_ko.get(ch, ch),
                        "캠페인": r.campaign.name,
                        "날짜": str(r.segments.date),
                        "기기": d_map.get(r.segments.device.name, "모바일"),
                        "노출수": int(r.metrics.impressions),
                        "클릭수": int(r.metrics.clicks),
                        "총비용": float(r.metrics.cost_micros) / 1_000_000,
                        "가입": float(r.metrics.conversions),
                    })
                    count += 1
            logs.append(f"✅ [Google] customer_id={cust_id} rows={count}")
        except Exception as e:
            logs.append(f"❌ [Google] customer_id={cust_id} 오류: {e}")

    return pd.DataFrame(rows), logs

def get_g_keyword_data(d_from, d_to):
    client = _google_client()
    ga = client.get_service("GoogleAdsService")

    query = f"""
        SELECT
          segments.date,
          segments.device,                -- ✅ 추가
          campaign.name,
          ad_group.name,
          ad_group_criterion.keyword.text,
          ad_group_criterion.keyword.match_type,
          metrics.impressions,
          metrics.clicks,
          metrics.cost_micros,
          metrics.conversions
        FROM keyword_view
        WHERE segments.date BETWEEN '{d_from}' AND '{d_to}'
          AND ad_group_criterion.type = KEYWORD
    """.strip()

    rows = []
    device_map = {"DESKTOP": "PC", "MOBILE": "모바일", "TABLET": "모바일", "CONNECTED_TV": "모바일"}

    # ✅ 하위 광고주 2개 모두 조회
    for cust_id in GOOGLE_CUSTOMER_IDS:
        try:
            stream = ga.search_stream(customer_id=cust_id, query=query)
            for b in stream:
                for r in b.results:
                    dev = r.segments.device.name
                    rows.append({
                        "매체": "구글",
                        "날짜": str(r.segments.date),
                        "캠페인": r.campaign.name,
                        "그룹": r.ad_group.name,
                        "키워드": r.ad_group_criterion.keyword.text,
                        "매칭": r.ad_group_criterion.keyword.match_type.name,
                        "기기": device_map.get(dev, "모바일"),
                        "노출 수": int(r.metrics.impressions),
                        "클릭 수": int(r.metrics.clicks),
                        "총 비용(VAT포함)": float(r.metrics.cost_micros) / 1_000_000,
                        "가입": float(r.metrics.conversions),
                        "평균노출순위": 0.0,
                    })
        except Exception as e:
            print(f"❌ [get_g_keyword_data] customer_id={cust_id} 오류: {e}")

    return pd.DataFrame(rows)

# =========================================================
# ✅ 6-2) 타뷸라 raw 파일 파싱
# =========================================================

# 깨진 캠페인명 → 정상 캠페인명 매핑
TABULA_CAMPAIGN_NAME_MAP = {
    "?щ갑???ㅼ씠?곕툕)_MO": "사방넷(네이티브)_MO",
    "?щ갑???ㅼ씠?곕툕)_PC": "사방넷(네이티브)_PC",
    "?щ갑??諛곕꼫)_MO":     "사방넷(배너)_MO",
    "?щ갑??諛곕꼫)_PC":     "사방넷(배너)_PC",
}

def parse_tabula_raw(file_path, logs=None) -> pd.DataFrame:
    """
    타뷸라 raw 파일(CSV or XLSX) 파싱
    컬럼: Campaign Name, Day, Impressions, Clicks, Spent, Conversions
    날짜 형식: DD-Mon-YY (예: 20-Feb-26)
    """
    if logs is None:
        logs = []

    if file_path is None:
        return pd.DataFrame()

    try:
        ext = str(file_path).lower()
        if ext.endswith(".xlsx") or ext.endswith(".xls"):
            df_raw = pd.read_excel(file_path)
        else:
            # CSV - 인코딩 자동 감지
            for enc in ["utf-8", "cp949", "euc-kr"]:
                try:
                    df_raw = pd.read_csv(file_path, encoding=enc)
                    break
                except Exception:
                    continue
            else:
                logs.append("❌ [타뷸라] CSV 인코딩 감지 실패")
                return pd.DataFrame()

        logs.append(f"[타뷸라] 원본 컬럼: {list(df_raw.columns)} rows={len(df_raw)}")

        # 컬럼명 정규화 (공백/대소문자)
        df_raw.columns = df_raw.columns.str.strip()

        # 필수 컬럼 확인
        required = {"Campaign Name", "Day", "Impressions", "Clicks", "Spent", "Conversions"}
        missing = required - set(df_raw.columns)
        if missing:
            logs.append(f"❌ [타뷸라] 필수 컬럼 없음: {missing}")
            return pd.DataFrame()

        df = df_raw.copy()

        # 깨진 캠페인명 복원
        df["Campaign Name"] = df["Campaign Name"].str.strip()
        df["Campaign Name"] = df["Campaign Name"].replace(TABULA_CAMPAIGN_NAME_MAP)
        logs.append(f"[타뷸라] 캠페인명 복원 완료: {df['Campaign Name'].unique().tolist()}")

        # 숫자 컬럼 쉼표 제거 후 변환
        for col in ["Impressions", "Clicks", "Spent", "Conversions"]:
            df[col] = df[col].astype(str).str.replace(",", "", regex=False)
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

        # 날짜 파싱 - 여러 형식 시도
        logs.append(f"[타뷸라] Day 샘플: {df['Day'].head(3).tolist()}")
        df["날짜"] = pd.to_datetime(df["Day"], format="%b %d, %Y", errors="coerce")
        if df["날짜"].isna().all():
            df["날짜"] = pd.to_datetime(df["Day"], format="%d-%b-%y", errors="coerce")
        if df["날짜"].isna().all():
            df["날짜"] = pd.to_datetime(df["Day"], errors="coerce")
        df["날짜"] = df["날짜"].dt.strftime("%Y-%m-%d")

        # 기기 추출 (_MO → 모바일, _PC → PC)
        df["기기"] = df["Campaign Name"].apply(
            lambda x: "PC" if str(x).endswith("_PC") else "모바일"
        )

        # 표준 컬럼으로 변환
        rows_df = pd.DataFrame({
            "매체구분": "DA",
            "매체": "타불라",
            "캠페인유형": "배너",
            "캠페인": df["Campaign Name"],
            "날짜": df["날짜"],
            "기기": df["기기"],
            "노출수": df["Impressions"].astype(int),
            "클릭수": df["Clicks"].astype(int),
            "총비용": df["Spent"],
            "가입": df["Conversions"],
        })

        rows_df = rows_df[rows_df["날짜"].notna()].reset_index(drop=True)
        logs.append(f"✅ [타뷸라] 파싱 완료 rows={len(rows_df)}")
        return rows_df

    except Exception as e:
        logs.append(f"❌ [타뷸라] 파싱 오류: {e}")
        return pd.DataFrame()

# =========================================================
# ✅ 7) 최종 통합리포트 만들기
# =========================================================
RAW_COLS = [
    "매체구분","매체","캠페인유형","캠페인","기간","기기",
    "요일","year","month","week","week시작","week종료",
    "노출수","클릭수","총비용","가입",
    "광고비(마크업포함,VAT포함)","서비스"
]

def build_final_df(platform: str, d_from: str, d_to: str, tabula_file=None):
    dfs = []
    logs = []

    if "Google" in platform:
        gdf, logs = get_g_data(d_from, d_to, logs)
        if not gdf.empty:
            dfs.append(gdf)

    if "Naver" in platform:
        ndf, logs = get_n_data(d_from, d_to, logs)
        if not ndf.empty:
            dfs.append(ndf)

    # ✅ 타뷸라 raw 파일 병합
    if tabula_file is not None:
        tdf = parse_tabula_raw(tabula_file, logs)
        if not tdf.empty:
            # 날짜 범위 필터링
            tdf_filtered = tdf[
                (tdf["날짜"] >= d_from) & (tdf["날짜"] <= d_to)
            ].reset_index(drop=True)
            logs.append(f"[타뷸라] 날짜 필터({d_from}~{d_to}) 후 rows={len(tdf_filtered)}")
            if not tdf_filtered.empty:
                dfs.append(tdf_filtered)

    if not dfs:
        return pd.DataFrame(columns=RAW_COLS), logs

    df = pd.concat(dfs, ignore_index=True)
    df = add_cal_fields(df, "날짜")
    df["서비스"] = assign_service_from_campaign(df["캠페인"])

    bs_keys = set(BS_DAILY_FEE_VAT_INCLUDED.keys())
    is_bs = df["캠페인"].isin(bs_keys)

    df.loc[~is_bs, "총비용"] = df.loc[~is_bs, "총비용"].apply(round_half_up_int)
    df.loc[~is_bs, "광고비(마크업포함,VAT포함)"] = df.loc[~is_bs].apply(calc_display_cost, axis=1)

    is_google = df["매체"].eq("구글")
    is_naver = df["매체"].eq("네이버")

    # ✅ FIX: BS 구글은 * 1.1, BS 네이버는 / 1.1 (VAT 이미 포함)
    df.loc[is_bs & is_google, "광고비(마크업포함,VAT포함)"] = (
        df.loc[is_bs & is_google, "총비용"].astype(float) * 1.1
    )
    df.loc[is_bs & is_naver, "광고비(마크업포함,VAT포함)"] = (
        df.loc[is_bs & is_naver, "총비용"].astype(float) / 1.1   # ✅ FIX: * → /
    )

    for c in RAW_COLS:
        if c not in df.columns:
            df[c] = ""

    df = df[RAW_COLS].sort_values("기간", na_position="last")
    return df, logs

def run_all(platform, d_f, d_t, tabula_file=None):
    try:
        logs = [f"APP_VERSION: {APP_VERSION}"]

        d_from = str(d_f)[:10]
        d_to = str(d_t)[:10]

        df, logs = build_final_df(platform, d_from, d_to, tabula_file)

        if df.empty:
            return "⚠️ 데이터가 없습니다.\n" + "\n".join(logs), None, None, platform

        fname = f"통합리포트_{datetime.now().strftime('%m%d_%H%M')}.xlsx"
        df.to_excel(fname, index=False)

        wb = load_workbook(fname)
        ws = wb.active
        header = [cell.value for cell in ws[1]]

        def col_idx(name):
            return header.index(name) + 1 if name in header else None

        col_period = col_idx("기간")
        col_week_s = col_idx("week시작")
        col_week_e = col_idx("week종료")
        col_cost = col_idx("총비용")
        col_cost2 = col_idx("광고비(마크업포함,VAT포함)")

        for r in range(2, ws.max_row + 1):
            if col_cost:
                ws.cell(row=r, column=col_cost).number_format = "#,##0"
            if col_cost2:
                ws.cell(row=r, column=col_cost2).number_format = "#,##0"
            for c in [col_period, col_week_s, col_week_e]:
                if c:
                    ws.cell(row=r, column=c).number_format = "yyyy-mm-dd"

        wb.save(fname)
        logs.append(f"✅ 최종 행수 = {len(df)}")
        return "\n".join(logs), fname, fname, platform

    except Exception:
        return f"❌ 오류:\n{traceback.format_exc()}", None, None, platform

# =========================================================
# ✅ 8) 키워드 성과 리포트 (Google + Naver)
# =========================================================
KW_FINAL_COLS = [
    "월","주간","매체","매체 구분","캠페인 유형","캠페인","그룹","키워드","기기",
    "노출 수","클릭 수","총 비용","가입","평균노출순위","가산","광고비(마크업포함,VAT포함)","서비스"
]

def _month_week_from_dt(dt_series: pd.Series):
    dt = pd.to_datetime(dt_series, errors="coerce")
    iso = dt.dt.isocalendar()
    month = dt.dt.year.astype("Int64").astype(str) + ". " + dt.dt.month.astype("Int64").astype(str).str.zfill(2)
    week = iso["week"].astype("Int64")
    return month, week

def infer_device_from_campaign_name_any(cname: str) -> str:
    s = str(cname or "")
    if s.endswith("_PC") or s.endswith("PC") or "_PC_" in s:
        return "PC"
    if s.endswith("_MO") or s.endswith("MO") or "_MO_" in s or "_M_" in s or "모바일" in s:
        return "모바일"
    return "전체"

NAVER_EXPKEYWORD_COLS = [
    "statDt", "customerId", "campaignId", "adgroupId", "keywordName",
    "bidAmt", "pcMblTp", "impCnt", "clkCnt", "ccnt", "salesAmt", "avgRnk"
]

def _parse_naver_expkeyword_txt(txt: str) -> pd.DataFrame:
    lines = [ln for ln in txt.splitlines() if str(ln).strip()]
    if not lines:
        return pd.DataFrame(columns=NAVER_EXPKEYWORD_COLS)

    df = pd.read_csv(
        io.StringIO("\n".join(lines)),
        sep="\t",
        header=None,
        names=NAVER_EXPKEYWORD_COLS,
        engine="python"
    )

    df["campaignName"] = df["campaignId"].map(camp_map).fillna(df["campaignId"])
    df["adgroupName"]  = df["adgroupId"].map(grp_map).fillna(df["adgroupId"])

    return df

def _naver_pc_mo_from_raw(pcMblTp: str) -> str:
    s = str(pcMblTp or "").upper().strip()
    if s in ("P", "PC"):
        return "PC"
    if s in ("M", "MOBILE", "MO"):
        return "모바일"
    return ""

def format_naver_keyword_report(nk_raw: pd.DataFrame) -> pd.DataFrame:
    nk = nk_raw.copy()

    for c in ["impCnt","clkCnt","ccnt","clkAmt","convAmt","avgRnk"]:
        if c in nk.columns:
            nk[c] = pd.to_numeric(nk[c], errors="coerce").fillna(0)

    dt = pd.to_datetime(nk["statDt"].astype(str).str[:8], format="%Y%m%d", errors="coerce")
    month_s, week_s = _month_week_from_dt(dt)

    out = pd.DataFrame()
    out["월"] = month_s
    out["주간"] = week_s
    out["매체"] = "네이버"
    out["매체 구분"] = "SA"
    out["캠페인 유형"] = "파워링크"
    out["캠페인"] = nk.get("campaignName", nk.get("campaignId", "")).astype(str)
    out["그룹"] = nk.get("adgroupName", nk.get("adgroupId", "")).astype(str)
    out["키워드"] = nk.get("keywordName", "").astype(str)

    out["기기"] = out["캠페인"].apply(infer_device_from_campaign_name_any)
    miss = out["기기"].isin(["", "전체"])
    if "pcMblTp" in nk.columns:
        out.loc[miss, "기기"] = nk.loc[miss, "pcMblTp"].apply(_naver_pc_mo_from_raw)

    out["노출 수"] = nk.get("impCnt", 0).astype(int)
    out["클릭 수"] = nk.get("clkCnt", 0).astype(int)

    # ✅ 광고비 컬럼 = clkAmt (클릭비용, 12번째 컬럼 = 실제 광고비)
    # convAmt(13번)는 전환매출액이라 광고비 아님! salesAmt 이름으로 잘못 매핑되어 있었음
    out["총 비용"] = pd.to_numeric(nk.get("clkAmt", 0), errors="coerce").fillna(0).apply(round_half_up_int)

    out["가입"] = pd.to_numeric(nk.get("ccnt", 0), errors="coerce").fillna(0).astype(int)
    out["평균노출순위"] = nk.get("avgRnk", 0).astype(float)

    out["가산"] = (out["노출 수"].astype(float) * out["평균노출순위"].astype(float)).fillna(0).round(1)
    # ✅ 네이버 clkAmt = VAT제외 → *1.1 해서 VAT포함 광고비 산출 (키워드 리포트는 clkAmt 기준이라 *1.1 유지)
    out["광고비(마크업포함,VAT포함)"] = out["총 비용"].apply(lambda x: round_half_up_int(float(x) * 1.1))
    out["서비스"] = assign_service_from_campaign(out["캠페인"].astype(str))

    for c in KW_FINAL_COLS:
        if c not in out.columns:
            out[c] = ""

    out = out[out["노출 수"].astype(int) > 0].reset_index(drop=True)

    return out[KW_FINAL_COLS]

def format_google_keyword_report(gk_raw: pd.DataFrame) -> pd.DataFrame:
    gk = gk_raw.copy()

    dt = pd.to_datetime(gk.get("날짜", ""), errors="coerce")
    month_s, week_s = _month_week_from_dt(dt)

    out = pd.DataFrame()
    out["월"] = month_s
    out["주간"] = week_s
    out["매체"] = "구글"
    out["매체 구분"] = "SA"
    out["캠페인 유형"] = "검색"

    out["캠페인"] = gk.get("캠페인", "").astype(str)
    out["그룹"]   = gk.get("그룹", "").astype(str)
    out["키워드"] = gk.get("키워드", "").astype(str)

    out["기기"] = gk.get("기기", "").astype(str)
    miss = out["기기"].isin(["", "전체"])
    out.loc[miss, "기기"] = out.loc[miss, "캠페인"].apply(infer_device_from_campaign_name_any)

    out["노출 수"] = pd.to_numeric(gk.get("노출 수", 0), errors="coerce").fillna(0).astype(int)
    out["클릭 수"] = pd.to_numeric(gk.get("클릭 수", 0), errors="coerce").fillna(0).astype(int)

    out["총 비용"] = pd.to_numeric(gk.get("총 비용(VAT포함)", 0), errors="coerce").fillna(0).apply(round_half_up_int)
    out["가입"] = pd.to_numeric(gk.get("가입", 0), errors="coerce").fillna(0).astype(float)

    out["평균노출순위"] = 0.0
    out["가산"] = (
        pd.to_numeric(out["노출 수"], errors="coerce").fillna(0) *
        pd.to_numeric(out["평균노출순위"], errors="coerce").fillna(0)
    ).round(1)

    out["광고비(마크업포함,VAT포함)"] = (out["총 비용"].astype(float) * 1.1).round(1)
    out["서비스"] = assign_service_from_campaign(out["캠페인"].astype(str))

    for c in KW_FINAL_COLS:
        if c not in out.columns:
            out[c] = ""

    out = out[out["노출 수"].astype(int) > 0].reset_index(drop=True)

    return out[KW_FINAL_COLS]


def _save_naver_raw_files(nk_raw: pd.DataFrame, prefix: str):
    if nk_raw is None or nk_raw.empty:
        return []

    ts = datetime.now().strftime("%m%d_%H%M")
    xlsx_path = f"{prefix}_{ts}.xlsx"
    csv_path  = f"{prefix}_{ts}.csv"

    nk_raw.to_excel(xlsx_path, index=False, engine="openpyxl")
    nk_raw.to_csv(csv_path, index=False, encoding="utf-8-sig")

    return [xlsx_path, csv_path]

def _pretty_rows(n: int) -> str:
    try:
        return f"{int(n):,}행"
    except Exception:
        return f"{n}행"

def run_keyword_report(platform, d1, d2):
    try:
        d_from = str(d1)[:10]
        d_to = str(d2)[:10]

        logs = [f"APP_VERSION: {APP_VERSION}"]
        out_dfs = []

        g_raw_n = 0
        n_raw_n = 0

        if "Google" in platform:
            gk_raw = get_g_keyword_data(d_from, d_to)
            g_raw_n = len(gk_raw)
            logs.append(f"Google keywords(raw): {g_raw_n}행")
            if not gk_raw.empty:
                gk_out = format_google_keyword_report(gk_raw)
                logs.append(f"Google keywords(formatted): {len(gk_out)}행")
                out_dfs.append(gk_out)

        if "Naver" in platform:
            if not NAVER_ACCOUNTS:
                logs.append("⚠️ NAVER 계정 환경변수 없음")
            else:
                NAVER_REPORT_TP_FOR_KEYWORD = os.environ.get("NAVER_KEYWORD_REPORT_TP", "AD")
                nk_raw = get_n_keyword_data_report(
                    d_from, d_to,
                    report_tp=NAVER_REPORT_TP_FOR_KEYWORD,
                    logs=logs
                )
                n_raw_n = len(nk_raw)
                logs.append(f"Naver keywords(raw): {n_raw_n}행 (reportTp={NAVER_REPORT_TP_FOR_KEYWORD})")

                if nk_raw.empty:
                    logs.append("⚠️ Naver report is empty")
                else:
                    logs.append(f"[진단] 컬럼: {list(nk_raw.columns)}")
                    logs.append(f"[진단] keywordName 있음: {'keywordName' in nk_raw.columns}")
                    if "keywordName" in nk_raw.columns:
                        logs.append(f"[진단] keywordName 비어있는 행: {nk_raw['keywordName'].eq('').sum()} / {len(nk_raw)}")
                        logs.append(f"[진단] unique 키워드 수: {nk_raw['keywordName'].nunique()}")
                    if "ccnt" in nk_raw.columns:
                        logs.append(f"[진단] ccnt>0 행: {(nk_raw['ccnt'] > 0).sum()} / {len(nk_raw)}")
                        logs.append(f"[진단] ccnt 합계: {int(nk_raw['ccnt'].sum())}")
                    nk_out = format_naver_keyword_report(nk_raw)
                    logs.append(f"Naver keywords(formatted): {len(nk_out)}행")
                    out_dfs.append(nk_out)

        if not out_dfs:
            summary = "⚠️ 키워드 데이터 없음"
            detail = "\n".join(logs)
            return summary, detail, None

        df_out = pd.concat(out_dfs, ignore_index=True)

        for c in KW_FINAL_COLS:
            if c not in df_out.columns:
                df_out[c] = ""
        df_out = df_out[KW_FINAL_COLS]

        fname = f"키워드성과_가공본_{datetime.now().strftime('%m%d_%H%M')}.xlsx"
        df_out.to_excel(fname, index=False)

        summary = (
            "✅ 키워드 성과 엑셀 생성 완료\n"
            f"- 구글: {_pretty_rows(g_raw_n)}\n"
            f"- 네이버: {_pretty_rows(n_raw_n)}\n"
            f"- 다운로드 파일 생성 완료"
        )
        detail = "\n".join(logs)

        return summary, detail, fname

    except Exception:
        return f"❌ 오류:\n{traceback.format_exc()}", "", None


# -----------------------------
# 데일리 코멘트 요약용 유틸
# -----------------------------

def _build_conv_keywords_map(platform: str, d: pd.Timestamp, top_n=5, logs=None, excel_path: str = None):
    if logs is None:
        logs = []

    d_from = d.strftime("%Y-%m-%d")
    d_to   = d.strftime("%Y-%m-%d")

    bucket = {}

    def _add(service, media, camp_type, group, keyword, conv):
        k = f"{service}|{media}|{camp_type}"
        if k not in bucket:
            bucket[k] = {}
        if group not in bucket[k]:
            bucket[k][group] = {}
        bucket[k][group][keyword] = bucket[k][group].get(keyword, 0) + int(conv)

    gk_df = pd.DataFrame()
    nk_df = pd.DataFrame()
    if excel_path:
        cache_path = excel_path.replace(".xlsx", "_kw_cache.json")
        if os.path.exists(cache_path):
            try:
                with open(cache_path, "r", encoding="utf-8") as f:
                    kw_cache = json.load(f)
                if "google" in kw_cache and kw_cache["google"]:
                    gk_df = pd.DataFrame(kw_cache["google"])
                    logs.append("✅ Google 키워드 캐시 사용")
                if "naver" in kw_cache and kw_cache["naver"]:
                    nk_df = pd.DataFrame(kw_cache["naver"])
                    logs.append("✅ Naver 키워드 캐시 사용")
            except Exception as e:
                logs.append(f"⚠️ 캐시 로드 실패, API 재호출: {e}")

    if "Google" in platform:
        try:
            gk = gk_df if not gk_df.empty else get_g_keyword_data(d_from, d_to)
            if not gk.empty:
                gk["서비스"] = assign_service_from_campaign(gk["캠페인"].astype(str))
                gk["가입"] = pd.to_numeric(gk.get("가입", 0), errors="coerce").fillna(0)
                gk = gk[gk["가입"] > 0]

                for _, r in gk.iterrows():
                    service = str(r.get("서비스", "")).strip()
                    media = "구글"
                    camp_type = "검색"
                    group = str(r.get("그룹", "")).strip() or "그룹"
                    keyword = str(r.get("키워드", "")).strip()
                    conv = int(r.get("가입", 0))
                    if service and keyword:
                        _add(service, media, camp_type, group, keyword, conv)
        except Exception as e:
            logs.append(f"⚠️ Google 키워드 전환 맵 실패: {e}")

    if "Naver" in platform:
        try:
            nk = nk_df if not nk_df.empty else get_n_keyword_data_report(d_from, d_to, report_tp="AD", logs=logs)
            if not nk.empty:
                nk["서비스"] = assign_service_from_campaign(nk.get("campaignName", "").astype(str))
                nk["가입"] = pd.to_numeric(nk.get("ccnt", 0), errors="coerce").fillna(0)
                nk = nk[nk["가입"] > 0]

                for _, r in nk.iterrows():
                    service = str(r.get("서비스", "")).strip()
                    media = "네이버"
                    camp_type = "파워링크"
                    group = str(r.get("adgroupName", "")).strip() or "그룹"
                    keyword = str(r.get("keywordName", "")).strip()
                    conv = int(r.get("가입", 0))
                    if service and keyword:
                        _add(service, media, camp_type, group, keyword, conv)
        except Exception as e:
            logs.append(f"⚠️ Naver 키워드 전환 맵 실패: {e}")

    out = {}
    for k, group_dict in bucket.items():
        group_totals = {g: sum(kw.values()) for g, kw in group_dict.items()}
        if not group_totals:
            continue

        best_group = sorted(group_totals.items(), key=lambda x: x[1], reverse=True)[0][0]
        kw_dict = group_dict.get(best_group, {})

        top_kw = sorted(kw_dict.items(), key=lambda x: x[1], reverse=True)[:top_n]
        out[k] = {
            "group": best_group,
            "keywords": [{"keyword": kw, "conv": int(cv)} for kw, cv in top_kw]
        }

    return out

def _pick_latest_date(df: pd.DataFrame) -> pd.Timestamp:
    col = "기간" if "기간" in df.columns else "날짜"
    dt = pd.to_datetime(df[col], errors="coerce")
    dt = dt.dropna()
    if dt.empty:
        raise ValueError("날짜/기간 컬럼에서 유효한 날짜를 찾지 못함")
    return dt.max().normalize()

def _filter_date(df: pd.DataFrame, target: pd.Timestamp) -> pd.DataFrame:
    col = "기간" if "기간" in df.columns else "날짜"
    dt = pd.to_datetime(df[col], errors="coerce").dt.normalize()
    return df.loc[dt == target.normalize()].copy()

def _agg(df: pd.DataFrame) -> dict:
    impr = pd.to_numeric(df.get("노출수", 0), errors="coerce").fillna(0).sum()
    clicks = pd.to_numeric(df.get("클릭수", 0), errors="coerce").fillna(0).sum()
    spend = pd.to_numeric(df.get("광고비(마크업포함,VAT포함)", 0), errors="coerce").fillna(0).sum()
    conv = pd.to_numeric(df.get("가입", 0), errors="coerce").fillna(0).sum()
    return {"impr": float(impr), "clicks": float(clicks), "spend": float(spend), "conv": float(conv)}

def _group(df: pd.DataFrame) -> pd.DataFrame:
    g = df.copy()
    for c in ["서비스", "매체", "캠페인유형"]:
        if c not in g.columns:
            g[c] = ""
    for c in ["노출수", "클릭수", "가입", "광고비(마크업포함,VAT포함)"]:
        if c not in g.columns:
            g[c] = 0

    g["노출수"] = pd.to_numeric(g["노출수"], errors="coerce").fillna(0)
    g["클릭수"] = pd.to_numeric(g["클릭수"], errors="coerce").fillna(0)
    g["가입"] = pd.to_numeric(g["가입"], errors="coerce").fillna(0)
    g["광고비(마크업포함,VAT포함)"] = pd.to_numeric(g["광고비(마크업포함,VAT포함)"], errors="coerce").fillna(0)

    out = g.groupby(["서비스", "매체", "캠페인유형"], dropna=False, as_index=False).agg(
        노출수=("노출수", "sum"),
        클릭수=("클릭수", "sum"),
        가입=("가입", "sum"),
        광고비=("광고비(마크업포함,VAT포함)", "sum"),
    )
    out = out.rename(columns={"광고비": "광고비(마크업포함,VAT포함)"})
    return out

def _safe_pct(cur: float, prev: float):
    try:
        cur = float(cur)
        prev = float(prev)
    except Exception:
        return None
    if prev == 0:
        return None
    return round((cur - prev) / prev * 100)

def _format_spend_delta_for_decrease(amount: float) -> str:
    try:
        amt = float(amount)
    except Exception:
        return ""
    if amt >= 0:
        return ""
    man = int(round(abs(amt) / 10000))
    if man <= 0:
        return ""
    return f"(-{man}만원)"

def _build_summary_for_ai(df_all: pd.DataFrame, compare_mode: str) -> dict:
    d = _pick_latest_date(df_all)
    p = _pick_compare_date(d, compare_mode)

    compare_rule = "전주 동요일 비교" if compare_mode == "전주 동요일(D-7) 비교" else "전일 비교"

    df_d = _filter_date(df_all, d)
    df_p = _filter_date(df_all, p)

    if df_p.empty:
        return {
            "ok": False,
            "message": f"비교일({p.strftime('%Y-%m-%d')}) 데이터가 파일에 없어 비교 코멘트 생성 불가",
        }

    a_d = _agg(df_d)
    a_p = _agg(df_p)

    gd = _group(df_d)
    gp = _group(df_p)

    m = gd.merge(
        gp,
        on=["서비스", "매체", "캠페인유형"],
        how="outer",
        suffixes=("_d", "_p")
    ).fillna(0)

    m["impr_pct"] = m.apply(lambda r: _safe_pct(float(r["노출수_d"]), float(r["노출수_p"])), axis=1)
    m["clicks_pct"] = m.apply(lambda r: _safe_pct(float(r["클릭수_d"]), float(r["클릭수_p"])), axis=1)
    m["spend_pct"] = m.apply(lambda r: _safe_pct(float(r["광고비(마크업포함,VAT포함)_d"]), float(r["광고비(마크업포함,VAT포함)_p"])), axis=1)
    m["conv_diff"] = (m["가입_d"] - m["가입_p"]).round().astype(int)

    m["spend_delta"] = (m["광고비(마크업포함,VAT포함)_d"] - m["광고비(마크업포함,VAT포함)_p"]).astype(float)
    m["abs_spend_delta"] = m["spend_delta"].abs()
    m["abs_conv_delta"] = m["conv_diff"].abs()

    top = m.sort_values(["abs_spend_delta", "abs_conv_delta"], ascending=False).head(6)
    service_order = {"사방넷": 0, "사방넷미니": 1, "풀필먼트": 2}
    top["_svc_order"] = top["서비스"].map(lambda x: service_order.get(str(x).strip(), 99))
    top = top.sort_values(["_svc_order", "abs_spend_delta", "abs_conv_delta"], ascending=[True, False, False])

    issues = []
    for _, r in top.iterrows():
        issues.append({
            "service": str(r["서비스"] or "").strip(),
            "media": str(r["매체"] or "").strip(),
            "campaign_type": str(r["캠페인유형"] or "").strip(),
            "impr_pct": None if pd.isna(r["impr_pct"]) else int(r["impr_pct"]) if r["impr_pct"] is not None else None,
            "clicks_pct": None if pd.isna(r["clicks_pct"]) else int(r["clicks_pct"]) if r["clicks_pct"] is not None else None,
            "spend_pct": None if pd.isna(r["spend_pct"]) else int(r["spend_pct"]) if r["spend_pct"] is not None else None,
            "conv_diff": int(r["conv_diff"]),
            "spend_delta": float(r["spend_delta"]),
        })

    return {
        "ok": True,
        "report_date": d.strftime("%Y-%m-%d"),
        "compare_date": p.strftime("%Y-%m-%d"),
        "compare_rule": compare_rule,
        "total": {
            "impr_pct": _safe_pct(a_d["impr"], a_p["impr"]),
            "clicks_pct": _safe_pct(a_d["clicks"], a_p["clicks"]),
            "spend_pct": _safe_pct(a_d["spend"], a_p["spend"]),
            "conv_diff": int(round(a_d["conv"] - a_p["conv"])),
        },
        "issues": issues,
        "rules": {
            "no_inference": True,
            "pct_only": True,
            "spend_decrease_amount_rule": "광고비 감소 시 증감액은 만원 단위까지만, 가능하면 생략",
        }
    }

# =========================================================
# ✅ 데일리 코멘트 생성
# =========================================================

ALLOWED_ENDINGS = ["증가", "감소", "발생", "확인", "예정", "영향"]

def _pick_compare_date(report_date: pd.Timestamp, compare_mode: str) -> pd.Timestamp:
    if compare_mode == "전주 동요일(D-7) 비교":
        return report_date - pd.Timedelta(days=7)
    return report_date - pd.Timedelta(days=1)


def generate_daily_comment_from_excel(excel_path: str, platform: str, compare_mode: str, manual_actions: str = "", include_kw: bool = False) -> str:
    df = pd.read_excel(excel_path)

    d = _pick_latest_date(df)
    p = _pick_compare_date(d, compare_mode)

    df_p = _filter_date(df, p)
    if df_p.empty:
        try:
            d_from = p.strftime("%Y-%m-%d")
            d_to = d.strftime("%Y-%m-%d")
            df, _logs = build_final_df(platform, d_from, d_to)
        except Exception:
            return (
                f"비교일({p.strftime('%Y-%m-%d')}) 데이터가 엑셀에 없고, "
                f"API로 보충도 실패\n{traceback.format_exc()}"
            )

    summary = _build_summary_for_ai(df, compare_mode)
    if not summary.get("ok"):
        return summary.get("message", "비교 데이터 없음")

    kw_map = {}
    if include_kw:
        kw_map = _build_conv_keywords_map(platform, d, top_n=5, excel_path=excel_path)

    for it in summary.get("issues", []):
        k = f"{it.get('service','')}|{it.get('media','')}|{it.get('campaign_type','')}"
        it["conv_kw_pack"] = kw_map.get(k, None)

    for it in summary["issues"]:
        it["spend_decrease_hint"] = _format_spend_delta_for_decrease(it.get("spend_delta", 0))

    instructions = f"""
너는 광고 일일 성과 보고용 데일리 코멘트 작성기다

[절대 규칙]
- 출력은 반드시 아래 구조로만 작성
  #주요이슈
  1. ...
  ㄴ ...
  2. ...
- 온점(.) 사용 금지
- 문장 끝은 {", ".join(ALLOWED_ENDINGS)} 중 하나로 종결
- "~하였습니다 / ~되었습니다" 금지
- 데이터에 없는 추측 해석 원인 생성 금지
- 인과관계 추정 금지 변동은 증감율(%) 중심으로만 작성
- 가입전환 증감 건수는 반드시 (+n건) (-n건) 형식으로 표기
- 광고비 감소 시 증감액 표기는 만원 단위까지만 가능하면 증감액 표기는 생략
- 하위 설명 'ㄴ'은 반드시 본문과 줄바꿈하여 별도 줄에 작성하며 들여쓰기 유지
- '전체' 또는 '전체 이슈' 문장 작성 금지
- 이슈는 항상 '사방넷'으로 시작 2번은 '사방넷미니' 3번은 '풀필먼트' 우선
ㄴ 줄에는 성과 비교 문장 작성 금지
- 번호 한 줄은 하나의 (서비스/매체/캠페인유형)

[키워드 규칙]
- 각 이슈에 conv_kw_packs가 있으면 반드시 ㄴ 줄을 1줄 작성
- ㄴ 줄은 아래 규칙으로 작성
  - 기본: '{{label}}에서 가입전환 {{conv}}건 발생' 형태를 여러 개 나열
  - 여러 개일 때는 쉼표로 연결
  - 예: ㄴ OW소재에서 가입전환 11건, 디지털보증서에서 4건 발생
- 단, label이 '경쟁사키워드' 또는 '브랜드 키워드' 등이고 keywords가 있으면 다음 형태 허용
  - ㄴ 경쟁사키워드 카페24, 셀메이트에서 각 1건씩 발생
  - ㄴ 브랜드 키워드에서 가입전환 9건 발생
- 키워드가 있는 경우에도 최대 3개까지만 표기
- conv_kw_packs가 없으면 키워드 ㄴ 줄은 작성 금지
- 키워드 ㄴ 줄이 있을 경우 manual_actions ㄴ 줄보다 먼저 작성


[conv_kw_pack 데이터 구조]
- conv_kw_pack 구조는 다음과 같다
  {{"group": "...", "keywords": [{{"keyword": "...", "conv": n}}, ...]}}

[번호 본문 작성 규칙]
- 각 번호 문장은 반드시 '전일대비' 또는 '전일에 이어' 같은 비교 표현으로 시작
- 본문에는 summary.issues의 impr_pct clicks_pct spend_pct conv_diff 중 의미 있는 값만 사용
- 원인 추정 금지
- 가입전환은 (+n건) (-n건) 형식 유지


[액션 메모 규칙]
- manual_actions가 비어있으면 액션 ㄴ 줄은 작성하지 말 것
- manual_actions가 있으면 그대로 복사하여 ㄴ 줄에만 사용
- 새로운 액션을 생성하거나 추측하지 말 것

""".strip()

    payload = {
        "variation_seed": summary["report_date"],
        "manual_actions": (manual_actions or "").strip(),
        "summary": summary,
    }

    full_prompt = instructions + "\n\n" + json.dumps(payload, ensure_ascii=False)
    resp = (_gemini.models.generate_content if _gemini else (_raise_no_key()))(
        model=GEMINI_MODEL,
        contents=full_prompt
    )

    text = (resp.text or "").strip()
    return text or "코멘트 생성 실패"


# =========================================================
# ✅ 9) UI  - Streamlit
# =========================================================

st.set_page_config(page_title="사방넷 리포트 도우미", page_icon="💖", layout="wide")

st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Pretendard:wght@400;600;700;800&display=swap');

:root, body, html {{
    font-family: 'Pretendard', system-ui, -apple-system, sans-serif !important;
    background: linear-gradient(180deg, #fff6fa 0%, #ffeef6 100%) !important;
}}

.hero {{
    border-radius: 28px;
    padding: 42px 48px;
    margin-bottom: 36px;
    background: linear-gradient(135deg, #ffd9ec, #ffeaf5);
    box-shadow:
        0 25px 50px rgba(255, 170, 200, 0.25),
        0 10px 20px rgba(255, 170, 200, 0.15);
    backdrop-filter: blur(6px);
}}

.hero h1 {{
    margin: 0;
    font-weight: 800;
    font-size: 34px;
    letter-spacing: -0.5px;
    color: #ff4fa3;
}}

.hero p {{
    margin-top: 10px;
    font-size: 15px;
    color: #c34c8f;
}}

.card {{
    border-radius: 26px;
    padding: 36px 40px;
    margin-bottom: 34px;
    background: #ffffff;
    border: 1px solid rgba(255, 192, 203, 0.25);
    box-shadow:
        0 20px 40px rgba(255, 182, 193, 0.18),
        0 8px 18px rgba(255, 182, 193, 0.12);
    transition: all 0.25s ease;
}}
.card:hover {{
    transform: translateY(-4px);
    box-shadow:
        0 28px 55px rgba(255, 160, 200, 0.28),
        0 12px 25px rgba(255, 160, 200, 0.18);
}}

button[kind="primary"] {{
    border-radius: 18px !important;
    background: linear-gradient(135deg, #ff9ecb, #ffb6dd) !important;
    font-weight: 800 !important;
    color: white !important;
    border: none !important;
    padding: 0.6rem 1.6rem !important;
    box-shadow:
        0 10px 20px rgba(255, 158, 203, 0.4),
        0 4px 10px rgba(255, 158, 203, 0.25) !important;
    transition: all 0.2s ease !important;
}}
button[kind="primary"]:hover {{
    transform: translateY(-2px);
    box-shadow:
        0 14px 28px rgba(255, 158, 203, 0.5),
        0 6px 14px rgba(255, 158, 203, 0.35) !important;
}}
button {{
    border-radius: 16px !important;
}}

button[role="tab"] {{
    font-weight: 700 !important;
    padding: 0.5rem 1rem !important;
}}
button[aria-selected="true"] {{
    color: #ff4fa3 !important;
    border-bottom: 3px solid #ff8dc5 !important;
}}

[data-testid="stDataFrame"] {{
    border-radius: 18px !important;
    overflow: hidden;
    box-shadow: 0 8px 20px rgba(255, 182, 193, 0.12);
}}

section[data-testid="stSidebar"] {{
    background: linear-gradient(180deg, #ffe9f4, #ffdff0);
}}
</style>

<div class="hero">
  <h1>💖 사방넷 리포트 도우미 💖</h1>
  <p>다우기술 사방넷 리포트 홈페이지 입니다 &gt;_&lt;</p>
</div>
""", unsafe_allow_html=True)

# =====================================================
# 대시보드 탭 (데일리 리포트 시각화)
# =====================================================
def render_daily_dashboard(df: pd.DataFrame, df_prev=None, d1=None, d2=None):
    NUM_COLS = ["노출수","클릭수","총비용","가입","광고비(마크업포함,VAT포함)"]
    for c in NUM_COLS:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    if df_prev is not None:
        for c in NUM_COLS:
            if c in df_prev.columns:
                df_prev[c] = pd.to_numeric(df_prev[c], errors="coerce").fillna(0)

    COST = "광고비(마크업포함,VAT포함)"

    def _s(d, col):
        return float(d[col].sum()) if d is not None and col in d.columns and len(d) > 0 else 0.0
    def _pct(a, b): return f"{a/b*100:.2f}%" if b > 0 else "-"
    def _cpc(cost, clk): return f"{int(cost/clk):,}" if clk > 0 else "-"
    def _cpa(cost, conv): return f"{int(cost/conv):,}원" if conv > 0 else "-"
    def _delta_badge(cur, prev):
        if prev == 0: return '<span style="color:#aaa;font-size:13px">-</span>'
        pct = (cur - prev) / prev * 100
        bg    = "rgba(220,252,231,0.9)" if pct >= 0 else "rgba(254,226,226,0.9)"
        color = "#15803d" if pct >= 0 else "#dc2626"
        arrow = "▲" if pct >= 0 else "▼"
        return f'<span style="background:{bg};color:{color};font-size:13px;font-weight:800;padding:2px 7px;border-radius:6px">{arrow}{abs(pct):.1f}%</span>'
    def _delta_sub(cur, prev, unit=""):
        if prev == 0: return f'<span style="color:rgba(255,255,255,0.4);font-size:11px">전일 {int(prev):,}{unit}</span>'
        pct = (cur - prev) / prev * 100
        color = "rgba(167,243,208,1)" if pct >= 0 else "rgba(252,165,165,1)"
        arrow = "▲" if pct >= 0 else "▼"
        return f'<span style="color:{color};font-size:14px;font-weight:800">{arrow}{abs(pct):.1f}%</span> <span style="color:rgba(255,255,255,0.4);font-size:11px">전일 {int(prev):,}{unit}</span>'

    period_label = f"{d1} ~ {d2}" if d1 and d2 else ""

    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;500;700;900&display=swap');

    .hero-kpi {
        background: linear-gradient(135deg, #be185d 0%, #db2777 40%, #ec4899 75%, #f9a8d4 100%);
        border-radius: 22px; padding: 26px 30px; color: white; margin-bottom: 18px;
        box-shadow: 0 8px 32px rgba(219,39,119,0.35);
    }
    .hero-title { font-size:17px; font-weight:900; letter-spacing:-0.3px; margin-bottom:18px;
                  display:flex; align-items:center; gap:10px; }
    .period-chip { background:rgba(255,255,255,0.2); backdrop-filter:blur(4px);
                   color:white; padding:3px 12px; border-radius:20px; font-size:11px; font-weight:600; }
    .hero-grid { display:grid; grid-template-columns:repeat(4,1fr); gap:0; }
    .hero-item { padding: 0 24px; border-left: 1px solid rgba(255,255,255,0.2); }
    .hero-item:first-child { border-left:none; padding-left:0; }
    .hi-label { font-size:10px; color:rgba(255,255,255,0.65); font-weight:600;
                letter-spacing:0.08em; text-transform:uppercase; margin-bottom:5px; }
    .hi-value { font-size:28px; font-weight:900; color:white; line-height:1.05; margin-bottom:4px; }
    .hi-unit  { font-size:13px; font-weight:500; color:rgba(255,255,255,0.65); }

    .svc-card { background:white; border-radius:18px; padding:22px 24px; margin-bottom:16px;
                box-shadow:0 4px 20px rgba(219,39,119,0.07); border:1px solid #fce7f3; }
    .svc-header { display:flex; align-items:center; gap:10px; margin-bottom:16px;
                  padding-bottom:12px; border-bottom:2px solid #fce7f3; }
    .svc-dot { width:10px; height:10px; border-radius:50%; flex-shrink:0; }
    .svc-title { font-size:16px; font-weight:900; color:#831843; }

    .perf-table { width:100%; border-collapse:separate; border-spacing:0; font-size:12px; }
    .perf-table thead tr { background:linear-gradient(90deg,#fdf2f8,#fce7f3); }
    .perf-table th { padding:10px 12px; font-size:10px; font-weight:700; color:#9d174d;
                     letter-spacing:0.06em; text-align:right; white-space:nowrap;
                     border-bottom:2px solid #fbcfe8; }
    .perf-table th:first-child { text-align:left; border-radius:10px 0 0 0; }
    .perf-table th:nth-child(2) { text-align:left; }
    .perf-table th:last-child { border-radius:0 10px 0 0; }
    .perf-table td { padding:11px 12px; text-align:right; color:#374151; font-weight:500;
                     border-bottom:1px solid #fdf2f8; white-space:nowrap; vertical-align:middle; }
    .perf-table td:first-child { text-align:left; }
    .perf-table td:nth-child(2) { text-align:left; color:#4b5563; font-weight:600; }
    .perf-table tbody tr:hover td { background:#fdf2f8; transition:background 0.15s; }
    .perf-table tfoot td { background:#fff0f7; font-weight:800; color:#831843;
                           border-top:2px solid #fbcfe8; border-bottom:none; padding:12px 12px; }
    .perf-table tfoot td:first-child { border-radius:0 0 0 10px; }
    .perf-table tfoot td:last-child { border-radius:0 0 10px 0; }

    .badge-n { display:inline-block; background:#dcfce7; color:#15803d; padding:3px 9px;
               border-radius:7px; font-size:10px; font-weight:800; }
    .badge-g { display:inline-block; background:#fee2e2; color:#dc2626; padding:3px 9px;
               border-radius:7px; font-size:10px; font-weight:800; }
    .badge-o { display:inline-block; background:#fff7ed; color:#c2410c; padding:3px 9px;
               border-radius:7px; font-size:10px; font-weight:800; }

    .conv-num { color:#9d174d; font-weight:900; font-size:14px; }
    .cost-num { color:#1d4ed8; font-weight:700; }
    .muted    { color:#9ca3af; }
    </style>
    """, unsafe_allow_html=True)

    tc = _s(df,"광고비(마크업포함,VAT포함)"); pc = _s(df_prev,"광고비(마크업포함,VAT포함)")
    ti = _s(df,"노출수"); pi = _s(df_prev,"노출수")
    tk = _s(df,"클릭수"); pk = _s(df_prev,"클릭수")
    tv = _s(df,"가입");   pv = _s(df_prev,"가입")

    st.markdown(f"""
    <div class="hero-kpi">
      <div class="hero-title">
        💖 광고 성과 대시보드
        <span class="period-chip">{period_label}</span>
        {"<span style='margin-left:auto;font-size:11px;color:rgba(255,255,255,0.55)'>🔄 전일 대비 비교 포함</span>" if df_prev is not None else ""}
      </div>
      <div class="hero-grid">
        <div class="hero-item">
          <div class="hi-label">총 광고비</div>
          <div class="hi-value">{int(tc):,}<span class="hi-unit">원</span></div>
          <div>{_delta_sub(tc,pc,"원")}</div>
        </div>
        <div class="hero-item">
          <div class="hi-label">총 노출수</div>
          <div class="hi-value">{int(ti):,}</div>
          <div>{_delta_sub(ti,pi)}</div>
        </div>
        <div class="hero-item">
          <div class="hi-label">총 클릭수</div>
          <div class="hi-value">{int(tk):,}</div>
          <div>{_delta_sub(tk,pk)}</div>
        </div>
        <div class="hero-item">
          <div class="hi-label">총 가입전환</div>
          <div class="hi-value" style="color:#fde68a">{int(tv):,}<span class="hi-unit">건</span></div>
          <div>{_delta_sub(tv,pv,"건")} <span style="color:rgba(255,255,255,0.4);font-size:11px">· CPA {_cpa(tc,tv)}</span></div>
        </div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    if "서비스" not in df.columns or "캠페인유형" not in df.columns:
        st.warning("서비스 또는 캠페인유형 컬럼이 없습니다.")
        return

    SVC_COLORS = {"사방넷":"#db2777","사방넷미니":"#0891b2","풀필먼트":"#059669"}
    MEDIA_BADGE = {
        "네이버":'<span class="badge-n">네이버</span>',
        "구글":  '<span class="badge-g">구글</span>',
        "타불라":'<span class="badge-o">타불라</span>',
        "타블라":'<span class="badge-o">타불라</span>',
    }

    services_order = ["사방넷","사방넷미니","풀필먼트"]
    services = [s for s in services_order if s in df["서비스"].unique()]
    extra = [s for s in df["서비스"].unique() if s not in services_order and str(s).strip() not in ("","nan","None")]
    services += extra

    for svc in services:
        svc_df   = df[df["서비스"]==svc]
        svc_prev = df_prev[df_prev["서비스"]==svc] if df_prev is not None and "서비스" in df_prev.columns else None

        sc = _s(svc_df,"광고비(마크업포함,VAT포함)"); psc = _s(svc_prev,"광고비(마크업포함,VAT포함)")
        si = _s(svc_df,"노출수");  sk = _s(svc_df,"클릭수"); sv = _s(svc_df,"가입"); psv = _s(svc_prev,"가입")

        dot = SVC_COLORS.get(svc,"#db2777")

        grp = svc_df.groupby(["매체","캠페인유형"], as_index=False).agg(
            노출=("노출수","sum"), 클릭=("클릭수","sum"),
            광고비=(COST,"sum"), 가입=("가입","sum")
        ).sort_values("광고비", ascending=False)

        if svc_prev is not None and "캠페인유형" in svc_prev.columns:
            gp = svc_prev.groupby(["매체","캠페인유형"], as_index=False).agg(
                광고비_p=(COST,"sum"), 가입_p=("가입","sum"))
            grp = grp.merge(gp, on=["매체","캠페인유형"], how="left").fillna({"광고비_p":0,"가입_p":0})
        else:
            grp["광고비_p"] = 0.0; grp["가입_p"] = 0.0

        rows = ""
        for _, r in grp.iterrows():
            badge = MEDIA_BADGE.get(str(r["매체"]), f'<span style="background:#f3f4f6;color:#374151;padding:3px 8px;border-radius:6px;font-size:10px;font-weight:700">{r["매체"]}</span>')
            conv_cell = f'<span class="conv-num">{int(r["가입"])}</span>' if r["가입"] > 0 else '<span class="muted">-</span>'
            db_cost = _delta_badge(r["광고비"], r["광고비_p"])
            db_conv = _delta_badge(r["가입"], r["가입_p"])
            rows += f"""<tr>
              <td>{badge}</td>
              <td>{r['캠페인유형']}</td>
              <td>{int(r['노출']):,}</td>
              <td>{int(r['클릭']):,}</td>
              <td class="muted">{_pct(r['클릭'],r['노출'])}</td>
              <td class="muted">{_cpc(r['광고비'],r['클릭'])}</td>
              <td><span class="cost-num">{int(r['광고비']):,}원</span>&nbsp;{db_cost}</td>
              <td>{conv_cell}&nbsp;{db_conv}</td>
              <td class="muted">{_pct(r['가입'],r['클릭'])}</td>
              <td class="muted">{_cpa(r['광고비'],r['가입'])}</td>
            </tr>"""

        db_sc = _delta_badge(sc, psc); db_sv = _delta_badge(sv, psv)
        foot = f"""<tr>
          <td colspan="2">합계</td>
          <td>{int(si):,}</td><td>{int(sk):,}</td>
          <td>{_pct(sk,si)}</td><td>{_cpc(sc,sk)}</td>
          <td>{int(sc):,}원&nbsp;{db_sc}</td>
          <td>{int(sv):,}건&nbsp;{db_sv}</td>
          <td>{_pct(sv,sk)}</td><td>{_cpa(sc,sv)}</td>
        </tr>"""

        st.markdown(f"""
        <div class="svc-card">
          <div class="svc-header">
            <span class="svc-dot" style="background:{dot}"></span>
            <span class="svc-title">{svc}</span>
            <span style="margin-left:auto;font-size:12px;color:#9d174d;font-weight:600">
              광고비 {int(sc):,}원 &nbsp;|&nbsp; 가입 {int(sv):,}건 &nbsp;|&nbsp; CPA {_cpa(sc,sv)}
            </span>
          </div>
          <table class="perf-table">
            <thead><tr>
              <th>매체</th><th>캠페인유형</th>
              <th>노출</th><th>클릭</th><th>CTR</th><th>CPC</th>
              <th>광고비 (전일비교)</th><th>가입 (전일비교)</th><th>가입율</th><th>CPA</th>
            </tr></thead>
            <tbody>{rows}</tbody>
            <tfoot>{foot}</tfoot>
          </table>
        </div>
        """, unsafe_allow_html=True)


# ── 세션 상태 초기화 ──────────────────────────────────────
if "saved_path"        not in st.session_state: st.session_state.saved_path        = None
if "saved_platform"    not in st.session_state: st.session_state.saved_platform    = None
if "chat_history"      not in st.session_state: st.session_state.chat_history      = []
if "daily_preset_prev" not in st.session_state: st.session_state.daily_preset_prev = "어제"
if "kw_preset_prev"    not in st.session_state: st.session_state.kw_preset_prev    = "주간(월~일)"
if "comment_out"       not in st.session_state: st.session_state.comment_out       = ""   # ✅ 추가

_r = preset_range("어제")
if "daily_d1" not in st.session_state: st.session_state.daily_d1 = datetime.strptime(_r[0][:10], "%Y-%m-%d").date()
if "daily_d2" not in st.session_state: st.session_state.daily_d2 = datetime.strptime(_r[1][:10], "%Y-%m-%d").date()

if "daily_log"      not in st.session_state: st.session_state.daily_log      = ""
if "daily_df"       not in st.session_state: st.session_state.daily_df       = None
if "daily_df_prev"  not in st.session_state: st.session_state.daily_df_prev  = None
if "daily_d1_saved" not in st.session_state: st.session_state.daily_d1_saved = None
if "daily_d2_saved" not in st.session_state: st.session_state.daily_d2_saved = None
if "daily_fname"    not in st.session_state: st.session_state.daily_fname    = None
if "kw_log"         not in st.session_state: st.session_state.kw_log         = ""
if "kw_detail_log"  not in st.session_state: st.session_state.kw_detail_log  = ""
if "kw_fname"       not in st.session_state: st.session_state.kw_fname       = None
if "kw_df"          not in st.session_state: st.session_state.kw_df          = None

# =====================================================
# 메인 레이아웃: 좌(대시보드) | 우(리포트+코멘트+챗봇)
# =====================================================
col_left, col_right = st.columns([6, 4])

# ══════════════════════════════════════════════════
# 왼쪽: 대시보드
# ══════════════════════════════════════════════════
with col_left:
    if st.session_state.daily_df is not None and not st.session_state.daily_df.empty:
        render_daily_dashboard(
            st.session_state.daily_df.copy(),
            st.session_state.daily_df_prev.copy() if st.session_state.daily_df_prev is not None else None,
            st.session_state.daily_d1_saved,
            st.session_state.daily_d2_saved,
        )
    else:
        st.markdown("""
        <div style="background:white;border-radius:20px;padding:40px;text-align:center;
                    border:2px dashed #e0d9ff;margin-top:20px">
          <div style="font-size:40px;margin-bottom:12px">📊</div>
          <div style="font-size:16px;font-weight:700;color:#7C6DEB;margin-bottom:6px">대시보드</div>
          <div style="font-size:13px;color:#9ca3af">오른쪽에서 데일리 리포트를 생성하면<br>여기에 성과가 표시됩니다</div>
        </div>
        """, unsafe_allow_html=True)

# ══════════════════════════════════════════════════
# 오른쪽: 리포트 생성 탭 + 코멘트 + 챗봇
# ══════════════════════════════════════════════════
with col_right:

    tab_daily, tab_kw = st.tabs(["📌 데일리 리포트", "🔎 키워드 성과"])

    with tab_daily:
        platform = st.radio("플랫폼", ["Google", "Naver", "Google+Naver"],
                            index=2, horizontal=True, key="daily_platform")

        preset = st.selectbox("기간", ["주간(월~일)", "어제", "지난 7일", "지난 30일", "이번 달", "직접선택"],
                              index=1, key="daily_preset")

        if preset != st.session_state.daily_preset_prev and preset != "직접선택":
            r = preset_range(preset)
            st.session_state.daily_d1 = datetime.strptime(r[0][:10], "%Y-%m-%d").date()
            st.session_state.daily_d2 = datetime.strptime(r[1][:10], "%Y-%m-%d").date()
            st.session_state.daily_preset_prev = preset
            st.rerun()

        col_d1, col_d2 = st.columns(2)
        with col_d1:
            d1 = st.date_input("시작일", key="daily_d1")
        with col_d2:
            d2 = st.date_input("종료일", key="daily_d2")

        tabula_file = st.file_uploader("📎 타뷸라 raw 파일 (선택)",
                                       type=["csv","xlsx"], key="tabula_upload")

        if st.button("통합 엑셀 생성", type="primary", key="btn_daily"):
            tabula_path = None
            if tabula_file:
                import tempfile
                suffix = ".xlsx" if tabula_file.name.endswith(".xlsx") else ".csv"
                with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                    tmp.write(tabula_file.read())
                    tabula_path = tmp.name

            with st.spinner("데이터 수집 중..."):
                log_msg, fname, saved, plat = run_all(
                    platform, str(d1), str(d2), tabula_path
                )
                try:
                    _df_today, _ = build_final_df(platform, str(d1), str(d2), tabula_path)
                    from datetime import timedelta as _td
                    _d1_prev = (datetime.strptime(str(d1), "%Y-%m-%d") - _td(days=1)).strftime("%Y-%m-%d")
                    _d2_prev = (datetime.strptime(str(d2), "%Y-%m-%d") - _td(days=1)).strftime("%Y-%m-%d")
                    _df_prev, _ = build_final_df(platform, _d1_prev, _d2_prev)
                    st.session_state.daily_df       = _df_today if not _df_today.empty else None
                    st.session_state.daily_df_prev  = _df_prev  if not _df_prev.empty  else None
                    st.session_state.daily_d1_saved = str(d1)
                    st.session_state.daily_d2_saved = str(d2)
                except Exception as _e:
                    st.session_state.daily_df = None
                    st.session_state.daily_df_prev = None

            st.session_state.daily_log = log_msg
            if fname and os.path.exists(fname):
                st.session_state.saved_path     = fname
                st.session_state.saved_platform = plat
                st.session_state.daily_fname    = fname

        if st.session_state.daily_log:
            with st.expander("📋 로그 보기"):
                st.text(st.session_state.daily_log)

        if st.session_state.daily_fname and os.path.exists(st.session_state.daily_fname):
            with open(st.session_state.daily_fname, "rb") as f:
                st.download_button("📥 통합 엑셀 다운로드", f,
                                   file_name=os.path.basename(st.session_state.daily_fname),
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   key="dl_daily")

    with tab_kw:
        kw_platform = st.radio("플랫폼", ["Google", "Naver", "Google+Naver"],
                               index=2, horizontal=True, key="kw_platform")

        kw_preset = st.selectbox("기간", ["주간(월~일)", "어제", "지난 7일", "지난 30일", "이번 달", "직접선택"],
                                 index=0, key="kw_preset")

        if kw_preset != st.session_state.kw_preset_prev and kw_preset != "직접선택":
            kr = preset_range(kw_preset)
            st.session_state.kw_d1 = datetime.strptime(kr[0][:10], "%Y-%m-%d").date()
            st.session_state.kw_d2 = datetime.strptime(kr[1][:10], "%Y-%m-%d").date()
            st.session_state.kw_preset_prev = kw_preset
            st.rerun()

        col_k1, col_k2 = st.columns(2)
        with col_k1:
            kw_d1 = st.date_input("시작일", key="kw_d1")
        with col_k2:
            kw_d2 = st.date_input("종료일", key="kw_d2")

        if st.button("키워드 성과 엑셀 생성", type="primary", key="btn_kw"):
            with st.spinner("키워드 데이터 수집 중..."):
                kw_summary, kw_detail, kw_fname = run_keyword_report(kw_platform, str(kw_d1), str(kw_d2))
            st.session_state.kw_log = kw_summary
            st.session_state.kw_detail_log = kw_detail
            if kw_fname and os.path.exists(kw_fname):
                st.session_state.kw_fname = kw_fname
                try:
                    st.session_state.kw_df = pd.read_excel(kw_fname)
                except Exception:
                    st.session_state.kw_df = None

        if st.session_state.kw_log:
            with st.expander("📋 로그 보기"):
                st.text(st.session_state.kw_log)
                if st.session_state.kw_detail_log:
                    st.text(st.session_state.kw_detail_log)

        if st.session_state.kw_fname and os.path.exists(st.session_state.kw_fname):
            with open(st.session_state.kw_fname, "rb") as f:
                st.download_button("📥 키워드 성과 다운로드", f,
                                   file_name=os.path.basename(st.session_state.kw_fname),
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   key="dl_kw")

    st.divider()

    # ── 하단: 코멘트 + 챗봇 탭 ───────────────────────
    tab_comment, tab_chat = st.tabs(["✅ 데일리 코멘트", "💬 챗봇"])

    with tab_comment:
        compare_mode = st.radio("비교 기준",
                                ["전일(D-1) 비교", "전주 동요일(D-7) 비교"],
                                horizontal=True, key="compare_mode")

        manual_actions = st.text_area("액션/메모 (옵션)",
                                      placeholder="예: 예산 상향 조정 후 모니터링 예정",
                                      height=60, key="manual_actions")

        include_kw = st.checkbox("키워드 인사이트 포함", value=False, key="include_kw_for_comment")

        if st.button("데일리 코멘트 생성", key="btn_comment", type="primary"):
            if not st.session_state.saved_path:
                st.warning("먼저 통합 엑셀을 생성해주세요")
            else:
                with st.spinner("코멘트 생성 중..."):
                    try:
                        result = generate_daily_comment_from_excel(
                            st.session_state.saved_path,
                            st.session_state.saved_platform,
                            compare_mode,
                            manual_actions or "",
                            include_kw=include_kw
                        )
                    except Exception:
                        result = f"❌ 코멘트 생성 오류:\n{traceback.format_exc()}"

                # ✅ 핵심 수정: session_state에 결과 저장
                st.session_state.comment_out = result

        # ✅ 핵심 수정: key 제거 → value만으로 표시 (session_state 충돌 없음)
        st.text_area("코멘트", value=st.session_state.comment_out, height=300)

    with tab_chat:
        # ── 챗봇 CSS ─────────────────────────────────────────
        st.markdown("""
        <style>
        .chat-wrap {
            display: flex; flex-direction: column; gap: 14px;
            padding: 16px 4px; max-height: 460px;
            overflow-y: auto;
        }
        .chat-row-user  { display:flex; justify-content:flex-end; align-items:flex-end; gap:8px; }
        .chat-row-ai    { display:flex; justify-content:flex-start; align-items:flex-end; gap:8px; }

        .chat-avatar-ai {
            width:34px; height:34px; border-radius:50%; flex-shrink:0;
            background: linear-gradient(135deg,#ff9ecb,#ffb6dd);
            display:flex; align-items:center; justify-content:center;
            font-size:16px; box-shadow:0 4px 10px rgba(255,158,203,0.4);
        }
        .chat-avatar-user {
            width:34px; height:34px; border-radius:50%; flex-shrink:0;
            background: linear-gradient(135deg,#a78bfa,#818cf8);
            display:flex; align-items:center; justify-content:center;
            font-size:15px; box-shadow:0 4px 10px rgba(167,139,250,0.35);
        }
        .bubble-ai {
            max-width:88%; padding:13px 16px; font-size:13px; line-height:1.65;
            background:#ffffff; color:#1f2937;
            border-radius:18px 18px 18px 4px;
            border:1px solid #fce7f3;
            box-shadow:0 4px 16px rgba(219,39,119,0.08);
            white-space: pre-wrap;
        }
        .bubble-user {
            max-width:82%; padding:11px 15px; font-size:13px; line-height:1.55;
            background: linear-gradient(135deg,#ec4899,#f472b6);
            color:white; font-weight:500;
            border-radius:18px 18px 4px 18px;
            box-shadow:0 4px 14px rgba(236,72,153,0.35);
        }
        .chat-empty {
            text-align:center; padding:40px 20px;
            color:#d1a0c0; font-size:13px;
        }
        .chat-empty-icon { font-size:36px; margin-bottom:8px; }
        .chat-empty-title { font-size:15px; font-weight:700; color:#f472b6; margin-bottom:4px; }

        .quick-chips { display:flex; flex-wrap:wrap; gap:7px; margin-bottom:12px; }
        .chip-label {
            display:inline-block; padding:5px 13px; font-size:11px; font-weight:600;
            border-radius:20px; cursor:pointer; border:1.5px solid #fbcfe8;
            color:#be185d; background:#fff0f7;
            transition:all 0.15s;
        }
        .chip-label:hover { background:#fbcfe8; }

        .chat-toolbar {
            display:flex; justify-content:space-between; align-items:center;
            margin-bottom:8px;
        }
        .chat-title-bar {
            font-size:13px; font-weight:800; color:#be185d;
            display:flex; align-items:center; gap:6px;
        }
        .online-dot {
            width:7px; height:7px; border-radius:50%;
            background:#22c55e; animation: blink 1.4s infinite;
        }
        @keyframes blink {
            0%,100%{ opacity:1; } 50%{ opacity:0.3; }
        }
        </style>
        """, unsafe_allow_html=True)

        # ── 툴바 ─────────────────────────────────────────────
        col_title, col_reset = st.columns([5, 1])
        with col_title:
            st.markdown("""
            <div class="chat-title-bar">
                <span class="online-dot"></span>
                💖 AI 광고 컨설턴트
                <span style="font-size:10px;color:#f9a8d4;font-weight:500;margin-left:4px">Powered by Gemini</span>
            </div>
            """, unsafe_allow_html=True)
        with col_reset:
            if st.button("🗑️", key="chat_reset", help="대화 초기화"):
                st.session_state.chat_history = []
                st.rerun()

        # ── 빠른 질문 칩 ─────────────────────────────────────
        QUICK_QUESTIONS = [
            "💰 CPA 개선 방법은?",
            "📈 가입전환 높은 캠페인 TOP3",
            "🔻 예산 낭비 캠페인 찾아줘",
            "🎯 지금 예산 배분 괜찮아?",
            "⚡ 지금 당장 할 수 있는 액션은?",
        ]

        cols = st.columns(len(QUICK_QUESTIONS))
        quick_clicked = None
        for i, (col, q) in enumerate(zip(cols, QUICK_QUESTIONS)):
            with col:
                if st.button(q, key=f"quick_{i}", use_container_width=True):
                    quick_clicked = q

        st.divider()

        # ── 채팅 히스토리 렌더링 ─────────────────────────────
        if not st.session_state.chat_history:
            st.markdown("""
            <div class="chat-empty">
                <div class="chat-empty-icon">💬</div>
                <div class="chat-empty-title">AI 광고 컨설턴트</div>
                데이터를 분석하고 솔루션을 드릴게요!<br>
                위 버튼을 눌러보거나 자유롭게 질문해주세요 ✨
            </div>
            """, unsafe_allow_html=True)
        else:
            chat_html = '<div class="chat-wrap">'
            for msg in st.session_state.chat_history:
                content = str(msg["content"]).replace("<", "&lt;").replace(">", "&gt;")
                if msg["role"] == "user":
                    chat_html += f"""
                    <div class="chat-row-user">
                        <div class="bubble-user">{content}</div>
                        <div class="chat-avatar-user">👤</div>
                    </div>"""
                else:
                    chat_html += f"""
                    <div class="chat-row-ai">
                        <div class="chat-avatar-ai">💖</div>
                        <div class="bubble-ai">{content}</div>
                    </div>"""
            chat_html += '</div>'
            st.markdown(chat_html, unsafe_allow_html=True)

        # ── 입력창 ────────────────────────────────────────────
        chat_input = st.chat_input("데이터 분석, 솔루션, 전략 뭐든 물어보세요!")
        final_input = quick_clicked or chat_input

        # ── 공통 Gemini 호출 로직 ─────────────────────────────
        def _build_data_context():
            """데이터 요약 + 원본 일부를 Gemini에게 전달"""
            if not st.session_state.saved_path:
                return "[데이터 없음 - 통합 엑셀을 먼저 생성해주세요]"
            try:
                df_ctx = pd.read_excel(st.session_state.saved_path)

                # 핵심 지표 요약
                summary_lines = ["=== 광고 성과 데이터 요약 ==="]

                # 전체 합산
                for col in ["노출수","클릭수","가입","광고비(마크업포함,VAT포함)","총비용"]:
                    if col in df_ctx.columns:
                        total = pd.to_numeric(df_ctx[col], errors="coerce").sum()
                        summary_lines.append(f"총 {col}: {total:,.0f}")

                # 날짜 범위
                if "기간" in df_ctx.columns:
                    dates = pd.to_datetime(df_ctx["기간"], errors="coerce").dropna()
                    if not dates.empty:
                        summary_lines.append(f"기간: {dates.min().date()} ~ {dates.max().date()}")

                # 서비스별 광고비 + 가입
                if "서비스" in df_ctx.columns and "광고비(마크업포함,VAT포함)" in df_ctx.columns:
                    svc_grp = df_ctx.groupby("서비스")[["광고비(마크업포함,VAT포함)","가입"]].sum()
                    summary_lines.append("\n=== 서비스별 성과 ===")
                    for svc, row in svc_grp.iterrows():
                        cost = row.get("광고비(마크업포함,VAT포함)", 0)
                        conv = row.get("가입", 0)
                        cpa = int(cost / conv) if conv > 0 else 0
                        summary_lines.append(f"{svc}: 광고비 {cost:,.0f}원 / 가입 {conv:.0f}건 / CPA {cpa:,}원")

                # 매체별 성과
                if "매체" in df_ctx.columns:
                    med_grp = df_ctx.groupby("매체")[["광고비(마크업포함,VAT포함)","가입","클릭수"]].sum()
                    summary_lines.append("\n=== 매체별 성과 ===")
                    for med, row in med_grp.iterrows():
                        cost = row.get("광고비(마크업포함,VAT포함)", 0)
                        conv = row.get("가입", 0)
                        clk  = row.get("클릭수", 0)
                        cpc  = int(cost / clk) if clk > 0 else 0
                        cpa  = int(cost / conv) if conv > 0 else 0
                        summary_lines.append(f"{med}: 광고비 {cost:,.0f}원 / 가입 {conv:.0f}건 / CPC {cpc:,}원 / CPA {cpa:,}원")

                # 캠페인별 성과 TOP 10 (광고비 기준)
                if "캠페인" in df_ctx.columns:
                    camp_grp = df_ctx.groupby("캠페인")[["광고비(마크업포함,VAT포함)","가입","클릭수","노출수"]].sum()
                    camp_grp = camp_grp.sort_values("광고비(마크업포함,VAT포함)", ascending=False).head(15)
                    summary_lines.append("\n=== 캠페인별 성과 TOP15 (광고비순) ===")
                    for camp, row in camp_grp.iterrows():
                        cost = row.get("광고비(마크업포함,VAT포함)", 0)
                        conv = row.get("가입", 0)
                        clk  = row.get("클릭수", 0)
                        imp  = row.get("노출수", 0)
                        ctr  = f"{clk/imp*100:.1f}%" if imp > 0 else "-"
                        cpa  = int(cost / conv) if conv > 0 else 0
                        summary_lines.append(
                            f"{camp}: 광고비 {cost:,.0f}원 / 가입 {conv:.0f}건 / CTR {ctr} / CPA {cpa:,}원"
                        )

                # raw 데이터 일부 (최근 100행)
                raw_sample = df_ctx.tail(100).to_string(index=False)
                return "\n".join(summary_lines) + "\n\n=== RAW 샘플 (최근 100행) ===\n" + raw_sample

            except Exception as e:
                return f"[데이터 로드 실패: {e}]"

        SYSTEM_PROMPT = """
너는 퍼포먼스 마케팅 전문가이자 데이터 기반 광고 컨설턴트야.
사방넷(셀링툴/풀필먼트/미니) 서비스의 광고 성과 데이터를 분석하고,
실질적인 솔루션과 액션 아이템을 제안하는 것이 주 임무야.

[역할 및 능력]
- 데이터에서 인사이트를 도출하고, 왜 그런 성과가 나왔는지 가설을 제시할 수 있어
- CPA, CTR, CPC, ROAS 등 광고 지표를 깊이 이해하고 벤치마크 대비 평가 가능
- 예산 재배분, 입찰 전략, 캠페인 구조 개선 등 구체적인 솔루션 제안 가능
- 네이버 파워링크, 구글 검색/디스플레이/PMax 각 플랫폼 특성을 알고 있어
- 브랜드검색(BS), 파워링크, 실적최대화 등 캠페인 유형별 전략 조언 가능

[답변 스타일]
- 친근하면서도 전문적으로, 마치 옆에 앉은 시니어 마케터처럼 대화해 다만 답변을 길게 말할 필요는 없어
- 데이터가 있으면 수치를 인용하면서 분석하고, 없으면 일반 마케팅 지식으로 답해
- 단순 데이터 조회 질문엔 수치를 깔끔하게 정리해서 보여줘
- 솔루션/전략 질문엔 우선순위 있는 액션 플랜으로 답해줘 (예: 📌 즉시 / 🔄 단기 / 📈 중장기)
- 숫자는 항상 쉼표(,) 단위로, 원화는 '원', 퍼센트는 '%' 명시
- 부정적인 성과도 솔직하게 말하고 개선 방향을 제시해줘
- 마크다운 형식(**, ##, - 등) 자유롭게 사용해서 가독성 높여줘

[주의사항]
- 데이터에 없는 구체적 수치는 지어내지 말고, "데이터에 없어서 정확한 수치는 모르지만~" 식으로 말해
- 하지만 마케팅 전문 지식을 기반으로 한 추론과 제안은 적극적으로 해도 돼
- 사방넷 서비스 특성(e-commerce 셀링툴, 풀필먼트, 미니샵)을 고려한 맞춤 조언을 해줘
""".strip()

        if final_input:
            st.session_state.chat_history.append({"role": "user", "content": final_input})

            with st.spinner("💭 분석 중..."):
                try:
                    data_ctx = _build_data_context()
                    full_system = SYSTEM_PROMPT + f"\n\n{data_ctx}"

                    gemini_history = [
                        {"role": "user",  "parts": [{"text": full_system}]},
                        {"role": "model", "parts": [{"text": "안녕하세요! 광고 데이터 확인했어요. 분석이나 솔루션이 필요한 게 있으면 편하게 물어보세요 💖"}]}
                    ]
                    for h in st.session_state.chat_history[:-1]:
                        role = "user" if h["role"] == "user" else "model"
                        gemini_history.append({"role": role, "parts": [{"text": h["content"]}]})
                    gemini_history.append({"role": "user", "parts": [{"text": final_input}]})

                    resp = (_gemini.models.generate_content if _gemini else (_raise_no_key()))(
                        model=GEMINI_MODEL, contents=gemini_history
                    )
                    answer = (resp.text or "").strip() or "응답 없음"
                except Exception as e:
                    answer = f"❌ 오류: {e}"

            st.session_state.chat_history.append({"role": "assistant", "content": answer})
            st.rerun()
